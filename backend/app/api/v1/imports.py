"""
Import management API endpoints.

Handles hotel and room data import from Excel/CSV files.
"""

import json
from datetime import datetime
from io import BytesIO
from typing import List, Optional

from fastapi import APIRouter, Depends, File, Path, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

import openpyxl
from app.core.database import get_db
from app.core.deps import get_current_user
from app.middleware.exception import BadRequestError, NotFoundError
from app.models.import_history import ImportStatus, ImportType
from app.models.user import User
from app.schemas.import_history import (
    ImportErrorsResponse,
    ImportErrorDetail,
    ImportHistoryBriefResponse,
    ImportHistoryListResponse,
    ImportHistoryResponse,
    ImportResultResponse,
    ImportResultRow,
)
from app.schemas.import_progress import ImportProgressResponse
from app.schemas.response import ApiResponse, PagedData, PagedResponse
from app.services.hotel_import_service import get_hotel_import_service
from app.services.room_import_service import get_room_import_service
from app.services.import_progress_service import get_import_progress_service

router = APIRouter()


def get_client_ip(request: Request) -> str:
    """Extract client IP from request."""
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    if request.client:
        return request.client.host
    return "unknown"


@router.post(
    "/hotels",
    response_model=ApiResponse[ImportHistoryResponse],
    status_code=201,
    summary="Import hotel data",
    description="Import hotel data from uploaded Excel or CSV file.",
)
async def import_hotels(
    request: Request,
    file: UploadFile = File(..., description="Excel or CSV file to import"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Import hotel data from Excel/CSV file.

    - **file**: Excel (.xlsx, .xls) or CSV file containing hotel data
    - Supported columns: hotel_id, name_cn, name_en, brand, status, country_code,
      province, city, district, address_cn, address_en, postal_code, phone, email,
      website, latitude, longitude, expedia_hotel_id, expedia_chain_code,
      expedia_property_code, opened_at, renovated_at
    """
    if not file.filename:
        raise BadRequestError(message="No file provided")

    # Validate file extension
    file_ext = file.filename.lower().split(".")[-1] if "." in file.filename else ""
    if file_ext not in {"xlsx", "xls", "csv"}:
        raise BadRequestError(
            message="Unsupported file format",
            details={"supported_formats": [".xlsx", ".xls", ".csv"]},
        )

    # Read file content
    content = await file.read()
    if not content:
        raise BadRequestError(message="Empty file provided")

    # Get operator info
    operator_id = current_user.id
    operator_name = current_user.full_name or current_user.username
    operator_ip = get_client_ip(request)

    # For now, use a temporary file path (in production, upload to cloud storage)
    file_path = f"imports/hotels/{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename}"

    # Import hotels
    service = get_hotel_import_service()
    import_history, row_results = await service.import_from_file(
        db=db,
        file_content=content,
        file_name=file.filename,
        file_path=file_path,
        operator_id=operator_id,
        operator_name=operator_name,
        operator_ip=operator_ip,
    )

    # Commit the transaction
    await db.commit()

    return ApiResponse(
        code=201,
        message=f"Hotel import completed: {import_history.success_rows} succeeded, {import_history.failed_rows} failed",
        data=ImportHistoryResponse.model_validate(import_history),
    )


@router.post(
    "/rooms",
    response_model=ApiResponse[ImportHistoryResponse],
    status_code=201,
    summary="Import room data",
    description="Import room data from uploaded Excel or CSV file.",
)
async def import_rooms(
    request: Request,
    file: UploadFile = File(..., description="Excel or CSV file to import"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Import room data from Excel/CSV file.

    - **file**: Excel (.xlsx, .xls) or CSV file containing room data
    - Supported columns: hotel_id, room_type_code, name_cn, name_en, description_cn,
      description_en, bed_type, max_occupancy, standard_occupancy, room_size,
      floor_range, total_rooms, expedia_room_id, expedia_room_type_code, is_active,
      amenities_cn, amenities_en, view_type, balcony, smoking_policy, bathroom_type
    """
    if not file.filename:
        raise BadRequestError(message="No file provided")

    # Validate file extension
    file_ext = file.filename.lower().split(".")[-1] if "." in file.filename else ""
    if file_ext not in {"xlsx", "xls", "csv"}:
        raise BadRequestError(
            message="Unsupported file format",
            details={"supported_formats": [".xlsx", ".xls", ".csv"]},
        )

    # Read file content
    content = await file.read()
    if not content:
        raise BadRequestError(message="Empty file provided")

    # Get operator info
    operator_id = current_user.id
    operator_name = current_user.full_name or current_user.username
    operator_ip = get_client_ip(request)

    # For now, use a temporary file path (in production, upload to cloud storage)
    file_path = f"imports/rooms/{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename}"

    # Import rooms
    service = get_room_import_service()
    import_history, row_results = await service.import_from_file(
        db=db,
        file_content=content,
        file_name=file.filename,
        file_path=file_path,
        operator_id=operator_id,
        operator_name=operator_name,
        operator_ip=operator_ip,
    )

    # Commit the transaction
    await db.commit()

    return ApiResponse(
        code=201,
        message=f"Room import completed: {import_history.success_rows} succeeded, {import_history.failed_rows} failed",
        data=ImportHistoryResponse.model_validate(import_history),
    )


@router.get(
    "",
    response_model=PagedResponse[ImportHistoryBriefResponse],
    summary="List import history",
    description="Get list of all import history records with pagination.",
)
async def list_imports(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Items per page"),
    import_type: Optional[ImportType] = Query(None, description="Filter by import type"),
    status: Optional[ImportStatus] = Query(None, description="Filter by status"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    List import history with pagination and filtering.

    - **page**: Page number (default: 1)
    - **page_size**: Items per page (default: 20, max: 100)
    - **import_type**: Filter by import type (hotel/room)
    - **status**: Filter by import status
    """
    # Get both hotel and room imports
    hotel_service = get_hotel_import_service()
    room_service = get_room_import_service()

    skip = (page - 1) * page_size
    limit = page_size

    # Query both services
    hotel_records, hotel_total = await hotel_service.get_import_history(db, skip=skip, limit=limit)
    room_records, room_total = await room_service.get_import_history(db, skip=skip, limit=limit)

    # Combine and sort by created_at
    all_records = hotel_records + room_records
    all_records.sort(key=lambda x: x.created_at, reverse=True)

    # Apply filters
    if import_type:
        all_records = [r for r in all_records if r.import_type == import_type]
    if status:
        all_records = [r for r in all_records if r.status == status]

    # Calculate total
    total = hotel_total + room_total
    if import_type:
        total = hotel_total if import_type == ImportType.HOTEL else room_total

    # Calculate pagination
    total_filtered = len(all_records)
    total_pages = (total_filtered + page_size - 1) // page_size if page_size > 0 else 0

    # Build response items
    items = []
    for record in all_records:
        item = ImportHistoryBriefResponse(
            id=record.id,
            file_name=record.file_name,
            import_type=record.import_type,
            status=record.status,
            total_rows=record.total_rows,
            success_rows=record.success_rows,
            failed_rows=record.failed_rows,
            skipped_rows=record.skipped_rows,
            success_rate=record.success_rate,
            started_at=record.started_at,
            completed_at=record.completed_at,
            processing_time=record.processing_time,
            operator_name=record.operator_name,
            created_at=record.created_at,
        )
        items.append(item)

    return PagedResponse(
        code=200,
        message="success",
        data=PagedData(
            list=items,
            total=total,
            page=page,
            page_size=page_size,
            total_pages=total_pages,
        ),
    )


@router.get(
    "/{import_id}",
    response_model=ApiResponse[ImportHistoryResponse],
    summary="Get import details",
    description="Get detailed information about a specific import.",
)
async def get_import(
    import_id: str = Path(..., description="Import history ID"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get import details by ID.

    - **import_id**: Import history ID
    """
    # Try hotel imports first
    hotel_service = get_hotel_import_service()
    record = await hotel_service.get_import_history_by_id(db, import_id)

    # If not found, try room imports
    if not record:
        room_service = get_room_import_service()
        record = await room_service.get_import_history_by_id(db, import_id)

    if not record:
        raise NotFoundError(
            message="Import not found",
            details={"import_id": import_id},
        )

    return ApiResponse(
        code=200,
        message="success",
        data=ImportHistoryResponse.model_validate(record),
    )


@router.get(
    "/{import_id}/errors",
    response_model=ApiResponse[ImportErrorsResponse],
    summary="Get import errors",
    description="Get detailed error information for a specific import.",
)
async def get_import_errors(
    import_id: str = Path(..., description="Import history ID"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get detailed errors for an import.

    - **import_id**: Import history ID
    """
    # Try hotel imports first
    hotel_service = get_hotel_import_service()
    record = await hotel_service.get_import_history_by_id(db, import_id)

    # If not found, try room imports
    if not record:
        room_service = get_room_import_service()
        record = await room_service.get_import_history_by_id(db, import_id)

    if not record:
        raise NotFoundError(
            message="Import not found",
            details={"import_id": import_id},
        )

    # Parse error log
    errors: List[ImportErrorDetail] = []
    if record.error_log:
        try:
            error_list = json.loads(record.error_log)
            for err in error_list:
                errors.append(ImportErrorDetail(
                    row=err.get("row"),
                    data=err.get("data"),
                    errors=err.get("errors", []),
                    message=err.get("message"),
                ))
        except json.JSONDecodeError:
            pass

    return ApiResponse(
        code=200,
        message="success",
        data=ImportErrorsResponse(
            import_id=record.id,
            file_name=record.file_name,
            total_errors=len(errors),
            errors=errors,
        ),
    )


@router.get(
    "/{import_id}/progress",
    response_model=ApiResponse[ImportProgressResponse],
    summary="Get import progress",
    description="Get real-time progress for an import operation.",
)
async def get_import_progress(
    import_id: str = Path(..., description="Import history ID"),
    current_user: User = Depends(get_current_user),
):
    """
    Get real-time progress for an import operation.

    - **import_id**: Import history ID
    """
    progress_service = get_import_progress_service()
    progress = await progress_service.get_progress(import_id)

    if not progress:
        raise NotFoundError(
            message="Import progress not found",
            details={"import_id": import_id},
        )

    return ApiResponse(
        code=200,
        message="success",
        data=ImportProgressResponse(**progress.to_dict()),
    )


@router.get(
    "/template/{import_type}",
    summary="Download import template",
    description="Download Excel template for hotel or room import.",
)
async def download_template(
    import_type: str = Path(..., description="Import type: 'hotels' or 'rooms'"),
):
    """
    Download import template Excel file.

    - **import_type**: 'hotels' for hotel template, 'rooms' for room template
    """
    if import_type not in ('hotels', 'rooms'):
        raise BadRequestError(
            message="Invalid import type. Must be 'hotels' or 'rooms'",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = import_type.capitalize()

    if import_type == 'hotels':
        headers = [
            'name_cn', 'name_en', 'brand', 'status', 'country_code', 'province', 'city',
            'district', 'address_cn', 'address_en', 'postal_code', 'phone', 'email',
            'latitude', 'longitude', 'check_in_time', 'check_out_time',
            'cancellation_policy', 'prepayment_policy', 'kid_policy', 'pet_policy',
            'services', 'facilities', 'description',
        ]
        # Add sample row
        sample_data = [
            '北京亚朵酒店', 'Atour Beijing', 'atour', 'draft', 'CN', '北京市',
            '北京', '朝阳区', '朝阳区某路123号', '123 Some Road, Chaoyang District', '100000',
            '+86-10-12345678', 'info@atour.com', 39.9042, 116.4074,
            '14:00', '12:00',
            '免费取消', '无需预付', '允许儿童入住', '不允许携带宠物',
            '免费WiFi,停车场', '健身房,会议室', '亚朵酒店为您提供舒适的住宿体验',
        ]
    else:
        headers = [
            'hotel_id', 'room_type_code', 'name_cn', 'name_en', 'description_cn',
            'bed_type', 'max_occupancy', 'standard_occupancy', 'room_size',
            'floor_range', 'total_rooms', 'amenities', 'smoking_policy',
        ]
        # Add sample row
        sample_data = [
            '{hotel_id}', 'STD-KING', '标准大床房', 'Standard King Room',
            '温馨舒适的标准大床房', 'King', 2, 2, 30.5,
            '3-10', 20, '免费WiFi,空调,电视', '禁止吸烟',
        ]

    ws.append(headers)
    ws.append(sample_data)

    # Format header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"import_{import_type}_template.xlsx"
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
