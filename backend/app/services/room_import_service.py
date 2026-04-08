"""
Room import service.

Handles room data import from Excel/CSV files with validation and tracking.
"""

import csv
import hashlib
import json
import os
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.hotel import Hotel, Room
from app.models.import_history import ImportHistory, ImportStatus, ImportType
from app.schemas.room import RoomCreate, RoomExtensionCreate
from app.services.room_service import RoomService, get_room_service
from app.validators.room_validator import RoomValidator, ValidationResult


class RoomImportService:
    """
    Service for room data import operations.

    Provides functionality to:
    - Import room data from Excel/CSV files
    - Validate room data before import
    - Track import history and errors
    """

    # Column name mappings (flexible parsing)
    COLUMN_MAPPINGS = {
        # Standard column names
        "hotel_id": ["hotel_id", "hotelid", "hotel"],
        "room_type_code": ["room_type_code", "roomtypecode", "room_code", "roomcode"],
        "name_cn": ["name_cn", "namecn", "room_name_cn", "roomnamecn"],
        "name_en": ["name_en", "nameen", "room_name_en", "roomnameen"],
        "description_cn": ["description_cn", "descriptioncn"],
        "description_en": ["description_en", "descriptionen"],
        "bed_type": ["bed_type", "bedtype"],
        "max_occupancy": ["max_occupancy", "maxoccupancy", "max_occ"],
        "standard_occupancy": ["standard_occupancy", "standardoccupancy", "std_occ"],
        "room_size": ["room_size", "roomsize", "size"],
        "floor_range": ["floor_range", "floorrange", "floor"],
        "total_rooms": ["total_rooms", "totalrooms", "rooms"],
        "expedia_room_id": ["expedia_room_id", "expediaroomid", "expedia_id"],
        "expedia_room_type_code": ["expedia_room_type_code", "expediaroomtypecode"],
        "is_active": ["is_active", "isactive", "active"],
        # Extension fields
        "amenities_cn": ["amenities_cn", "amenitiescn"],
        "amenities_en": ["amenities_en", "amenitiesen"],
        "view_type": ["view_type", "viewtype"],
        "balcony": ["balcony"],
        "smoking_policy": ["smoking_policy", "smokingpolicy"],
        "bathroom_type": ["bathroom_type", "bathroomtype"],
    }

    def __init__(self, room_service: Optional[RoomService] = None):
        """
        Initialize RoomImportService.

        Args:
            room_service: Optional RoomService instance. Defaults to global instance.
        """
        self.room_service = room_service or get_room_service()
        self.validator = RoomValidator()

    def _find_column(self, headers: List[str], field: str) -> Optional[int]:
        """
        Find column index by field name (case-insensitive).

        Args:
            headers: List of column headers
            field: Field name to find

        Returns:
            Column index or None if not found
        """
        field_lower = field.lower()
        possible_names = self.COLUMN_MAPPINGS.get(field_lower, [field_lower])

        for i, header in enumerate(headers):
            header_lower = header.lower().strip()
            if header_lower in possible_names:
                return i

        return None

    def _normalize_boolean(self, value: Any) -> Optional[bool]:
        """
        Normalize boolean values from various representations.

        Args:
            value: Value to normalize

        Returns:
            Boolean value or None
        """
        if value is None:
            return None

        if isinstance(value, bool):
            return value

        if isinstance(value, (int, float)):
            return bool(value)

        if isinstance(value, str):
            value_lower = value.lower().strip()
            if value_lower in ("true", "yes", "1", "t", "y"):
                return True
            if value_lower in ("false", "no", "0", "f", "n"):
                return False

        return None

    def _normalize_value(self, value: Any, field: str) -> Any:
        """
        Normalize value based on field type.

        Args:
            value: Value to normalize
            field: Field name

        Returns:
            Normalized value
        """
        if value is None or value == "":
            return None

        # Boolean fields
        if field in ("is_active", "balcony"):
            return self._normalize_boolean(value)

        # Integer fields
        if field in ("max_occupancy", "standard_occupancy", "total_rooms"):
            try:
                return int(float(value))
            except (ValueError, TypeError):
                return None

        # Float fields
        if field in ("room_size",):
            try:
                return float(value)
            except (ValueError, TypeError):
                return None

        # String fields - strip whitespace
        if isinstance(value, str):
            return value.strip()

        return value

    def _parse_headers(self, headers: List[str]) -> Dict[str, int]:
        """
        Parse headers and create field-to-index mapping.

        Args:
            headers: List of column headers

        Returns:
            Dictionary mapping field names to column indices
        """
        field_map: Dict[str, int] = {}

        for field in self.COLUMN_MAPPINGS.keys():
            col_idx = self._find_column(headers, field)
            if col_idx is not None:
                field_map[field] = col_idx

        return field_map

    def _row_to_dict(self, row: List[str], field_map: Dict[str, int]) -> Dict[str, Any]:
        """
        Convert a CSV row to a dictionary using field map.

        Args:
            row: List of cell values
            field_map: Mapping of field names to column indices

        Returns:
            Dictionary of field names to values
        """
        result: Dict[str, Any] = {}

        for field, col_idx in field_map.items():
            if col_idx < len(row):
                value = row[col_idx]
                result[field] = self._normalize_value(value, field)

        return result

    def _parse_excel_file(self, content: BytesIO) -> List[Dict[str, Any]]:
        """
        Parse Excel file content.

        Args:
            content: File content as BytesIO

        Returns:
            List of row dictionaries
        """
        if openpyxl is None:
            raise ImportError("openpyxl is required for Excel file parsing")

        wb = openpyxl.load_workbook(content)
        ws = wb.active

        # Read headers from first row
        headers = [cell.value for cell in ws[1]]
        if not headers:
            raise ValueError("Excel file has no headers")

        # Parse headers
        field_map = self._parse_headers([str(h) if h else "" for h in headers])

        if not field_map:
            raise ValueError("No recognized room fields found in Excel headers")

        # Read data rows
        rows: List[Dict[str, Any]] = []
        for row_idx in range(2, ws.max_row + 1):
            row_values = [ws.cell(row=row_idx, column=col_idx + 1).value
                          for col_idx in range(len(headers))]
            row_dict = self._row_to_dict(
                [str(v) if v is not None else "" for v in row_values],
                field_map
            )
            rows.append(row_dict)

        return rows

    def _parse_csv_file(self, content: str) -> List[Dict[str, Any]]:
        """
        Parse CSV file content.

        Args:
            content: File content as string

        Returns:
            List of row dictionaries
        """
        reader = csv.reader(StringIO(content))
        headers = next(reader)

        if not headers:
            raise ValueError("CSV file has no headers")

        # Parse headers
        field_map = self._parse_headers(headers)

        if not field_map:
            raise ValueError("No recognized room fields found in CSV headers")

        # Read data rows
        rows: List[Dict[str, Any]] = []
        for row in reader:
            row_dict = self._row_to_dict(row, field_map)
            rows.append(row_dict)

        return rows

    def _compute_file_hash(self, content: bytes) -> str:
        """
        Compute MD5 hash of file content.

        Args:
            content: File content bytes

        Returns:
            MD5 hash string
        """
        return hashlib.md5(content).hexdigest()

    async def validate_before_import(
        self,
        db: AsyncSession,
        room_data: Dict[str, Any],
    ) -> ValidationResult:
        """
        Validate room data before import.

        Args:
            db: Database session
            room_data: Room data dictionary

        Returns:
            ValidationResult with validation status and errors
        """
        result = self.validator.validate_room_data(room_data)

        # Additional database checks
        hotel_id = room_data.get("hotel_id")
        if hotel_id:
            hotel_result = await db.execute(select(Hotel).where(Hotel.id == hotel_id))
            if hotel_result.scalar_one_or_none() is None:
                result.add_error("hotel_id", f"Hotel with ID {hotel_id} does not exist", hotel_id)

        return result

    async def import_single(
        self,
        db: AsyncSession,
        room_data: Dict[str, Any],
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
    ) -> Tuple[Optional[Room], ValidationResult]:
        """
        Import a single room.

        Args:
            db: Database session
            room_data: Room data dictionary
            operator_id: Optional operator ID for audit
            operator_name: Optional operator name for audit

        Returns:
            Tuple of (created Room or None, ValidationResult)
        """
        # Validate
        validation_result = await self.validate_before_import(db, room_data)

        if not validation_result.is_valid:
            return None, validation_result

        try:
            # Extract extension data
            extension_fields = {
                "amenities_cn", "amenities_en", "amenity_details", "image_urls",
                "thumbnail_url", "view_type", "balcony", "smoking_policy",
                "floor", "bathroom_type", "bathroom_amenities_cn",
                "bathroom_amenities_en", "accessibility_features",
            }
            room_fields = {k: v for k, v in room_data.items() if k not in extension_fields}
            extension_data = {k: v for k, v in room_data.items() if k in extension_fields and v is not None}

            # Create room
            room_create = RoomCreate(**room_data)

            # Create extension if has extension data
            extension_create = None
            if extension_data and "room_id" not in extension_data:
                # We need to create room first to get the ID
                pass

            room = await self.room_service.create_room(
                db,
                room_in=room_create,
                extension_in=None,  # Will handle extension separately after room creation
            )

            # Create extension if we have extension data
            if extension_data:
                extension_create = RoomExtensionCreate(room_id=room.id, **extension_data)
                await self.room_service.update_room_extension(
                    db,
                    room_id=room.id,
                    extension_in=extension_create,
                )

            return room, validation_result

        except Exception as e:
            validation_result.add_error("_system", f"Import failed: {str(e)}")
            return None, validation_result

    async def import_from_file(
        self,
        db: AsyncSession,
        file_content: bytes,
        file_name: str,
        file_path: str,
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
        operator_ip: Optional[str] = None,
    ) -> Tuple[ImportHistory, List[Dict[str, Any]]]:
        """
        Import room data from Excel/CSV file.

        Args:
            db: Database session
            file_content: File content bytes
            file_name: Original file name
            file_path: Storage path for the file
            operator_id: Optional operator ID for audit
            operator_name: Optional operator name for audit
            operator_ip: Optional operator IP for audit

        Returns:
            Tuple of (ImportHistory record, list of row results)
        """
        # Compute file hash
        file_hash = self._compute_file_hash(file_content)
        file_size = len(file_content)

        # Determine file type
        is_excel = file_name.lower().endswith((".xlsx", ".xls"))
        is_csv = file_name.lower().endswith(".csv")

        if not is_excel and not is_csv:
            raise ValueError("Unsupported file format. Please use Excel or CSV file.")

        # Create import history record
        import_history = ImportHistory(
            file_name=file_name,
            file_path=file_path,
            file_size=file_size,
            file_hash=file_hash,
            import_type=ImportType.ROOM,
            status=ImportStatus.PROCESSING,
            total_rows=0,
            success_rows=0,
            failed_rows=0,
            skipped_rows=0,
            operator_id=operator_id,
            operator_name=operator_name,
            operator_ip=operator_ip,
            started_at=datetime.utcnow(),
        )
        db.add(import_history)
        await db.flush()

        errors: List[Dict[str, Any]] = []
        warnings: List[Dict[str, Any]] = []
        row_results: List[Dict[str, Any]] = []
        start_time = datetime.utcnow()

        try:
            # Parse file
            if is_excel:
                rows = self._parse_excel_file(BytesIO(file_content))
            else:
                content = file_content.decode("utf-8")
                rows = self._parse_csv_file(content)

            total_rows = len(rows)
            import_history.total_rows = total_rows

            # Process each row
            for idx, row_data in enumerate(rows):
                row_num = idx + 2  # Excel row number (1-indexed, header is row 1)
                row_result: Dict[str, Any] = {
                    "row": row_num,
                    "success": False,
                    "errors": [],
                    "warnings": [],
                }

                try:
                    # Validate
                    validation_result = await self.validate_before_import(db, row_data)

                    if not validation_result.is_valid:
                        row_result["errors"] = validation_result.to_dict()["errors"]
                        errors.append({
                            "row": row_num,
                            "data": row_data,
                            "errors": validation_result.to_dict()["errors"],
                        })
                        import_history.failed_rows += 1
                    else:
                        # Import room
                        room, _ = await self.import_single(
                            db,
                            row_data,
                            operator_id=operator_id,
                            operator_name=operator_name,
                        )

                        if room:
                            row_result["success"] = True
                            row_result["room_id"] = room.id
                            import_history.success_rows += 1
                        else:
                            import_history.failed_rows += 1

                except Exception as e:
                    row_result["errors"] = [{"field": "_system", "message": str(e)}]
                    errors.append({
                        "row": row_num,
                        "data": row_data,
                        "errors": [{"field": "_system", "message": str(e)}],
                    })
                    import_history.failed_rows += 1

                row_results.append(row_result)

            # Update status
            if import_history.failed_rows == 0:
                import_history.status = ImportStatus.COMPLETED
            elif import_history.success_rows == 0:
                import_history.status = ImportStatus.FAILED
            else:
                import_history.status = ImportStatus.PARTIAL

            # Calculate processing time
            end_time = datetime.utcnow()
            import_history.completed_at = end_time
            import_history.processing_time = (end_time - start_time).total_seconds()

            # Store error and warning logs
            if errors:
                import_history.error_log = json.dumps(errors, ensure_ascii=False)
            if warnings:
                import_history.warning_log = json.dumps(warnings, ensure_ascii=False)

            await db.flush()

        except Exception as e:
            import_history.status = ImportStatus.FAILED
            import_history.error_log = json.dumps([{"row": "file", "message": str(e)}], ensure_ascii=False)
            import_history.completed_at = datetime.utcnow()
            import_history.processing_time = (datetime.utcnow() - start_time).total_seconds()
            await db.flush()
            raise

        return import_history, row_results

    async def get_import_history(
        self,
        db: AsyncSession,
        skip: int = 0,
        limit: int = 20,
    ) -> Tuple[List[ImportHistory], int]:
        """
        Get import history records.

        Args:
            db: Database session
            skip: Number of records to skip
            limit: Maximum number of records to return

        Returns:
            Tuple of (list of ImportHistory, total count)
        """
        result = await db.execute(
            select(ImportHistory)
            .where(ImportHistory.import_type == ImportType.ROOM)
            .order_by(ImportHistory.created_at.desc())
            .offset(skip)
            .limit(limit)
        )
        records = list(result.scalars().all())

        count_result = await db.execute(
            select(ImportHistory)
            .where(ImportHistory.import_type == ImportType.ROOM)
        )
        total = len(list(count_result.scalars().all()))

        return records, total

    async def get_import_history_by_id(
        self,
        db: AsyncSession,
        import_id: str,
    ) -> Optional[ImportHistory]:
        """
        Get a specific import history record.

        Args:
            db: Database session
            import_id: Import history ID

        Returns:
            ImportHistory instance or None
        """
        result = await db.execute(
            select(ImportHistory).where(ImportHistory.id == import_id)
        )
        return result.scalar_one_or_none()


# Singleton instance
_room_import_service: Optional[RoomImportService] = None


def get_room_import_service() -> RoomImportService:
    """Get room import service singleton."""
    global _room_import_service
    if _room_import_service is None:
        _room_import_service = RoomImportService()
    return _room_import_service
