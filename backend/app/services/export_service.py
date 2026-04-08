"""
Export service for handling hotel and room data exports.
"""

import csv
import json
import os
import time
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from io import BytesIO
from openpyxl import Workbook
from sqlalchemy import and_, select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.config import settings
from app.models.export_history import ExportFormat, ExportHistory, ExportStatus, ExportType
from app.models.hotel import Hotel, Room
from app.models.room import RoomExtension
from app.generators.excel_template_generator import get_excel_template_generator


class ExportService:
    """
    Service for exporting hotel and room data to various formats.
    """

    def __init__(self):
        """Initialize the export service."""
        self.export_dir = os.path.join(settings.UPLOAD_DIR, "exports")
        os.makedirs(self.export_dir, exist_ok=True)

    async def create_export_record(
        self,
        db: AsyncSession,
        *,
        export_type: ExportType,
        export_format: ExportFormat,
        hotel_ids: Optional[List[str]] = None,
        use_template: bool = False,
        template_id: Optional[str] = None,
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
        operator_ip: Optional[str] = None,
    ) -> ExportHistory:
        """
        Create a new export record.

        Args:
            db: Database session
            export_type: Type of export (hotel/room)
            export_format: Export format (excel/csv/json)
            hotel_ids: Optional list of hotel IDs to filter
            use_template: Whether to use Expedia template
            template_id: Template ID if using template
            operator_id: User ID who initiated export
            operator_name: User name who initiated export
            operator_ip: IP address of the user

        Returns:
            Created ExportHistory instance
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{export_type.value}_{export_format.value}_{timestamp}"

        export_record = ExportHistory(
            file_name=file_name,
            export_type=export_type,
            export_format=export_format,
            status=ExportStatus.PENDING,
            hotel_ids=json.dumps(hotel_ids) if hotel_ids else None,
            filter_criteria=json.dumps({"use_template": use_template}),
            template_id=template_id,
            operator_id=operator_id,
            operator_name=operator_name,
            operator_ip=operator_ip,
            started_at=datetime.now(),
        )
        db.add(export_record)
        await db.flush()
        await db.refresh(export_record)
        return export_record

    async def get_export_record(
        self, db: AsyncSession, *, export_id: str
    ) -> Optional[ExportHistory]:
        """
        Get an export record by ID.

        Args:
            db: Database session
            export_id: Export record ID

        Returns:
            ExportHistory instance or None
        """
        result = await db.execute(
            select(ExportHistory).where(ExportHistory.id == export_id)
        )
        return result.scalar_one_or_none()

    async def update_export_status(
        self,
        db: AsyncSession,
        *,
        export_id: str,
        status: ExportStatus,
        file_path: Optional[str] = None,
        file_size: Optional[int] = None,
        download_url: Optional[str] = None,
        total_rows: int = 0,
        total_hotels: int = 0,
        total_rooms: int = 0,
        error_message: Optional[str] = None,
        completed: bool = False,
    ) -> Optional[ExportHistory]:
        """
        Update export record status.

        Args:
            db: Database session
            export_id: Export record ID
            status: New status
            file_path: Path to generated file
            file_size: Size of generated file
            download_url: Download URL
            total_rows: Total rows exported
            total_hotels: Total hotels exported
            total_rooms: Total rooms exported
            error_message: Error message if failed
            completed: Whether export is completed

        Returns:
            Updated ExportHistory instance or None
        """
        export_record = await self.get_export_record(db, export_id=export_id)
        if not export_record:
            return None

        export_record.status = status
        if file_path:
            export_record.file_path = file_path
        if file_size is not None:
            export_record.file_size = file_size
        if download_url:
            export_record.download_url = download_url
        export_record.total_rows = total_rows
        export_record.total_hotels = total_hotels
        export_record.total_rooms = total_rooms
        if error_message:
            export_record.error_message = error_message

        if completed:
            export_record.completed_at = datetime.now()
            export_record.processing_time = (
                export_record.completed_at - export_record.started_at
            ).total_seconds() if export_record.started_at else None
            export_record.expires_at = datetime.now() + timedelta(days=7)

        db.add(export_record)
        await db.flush()
        await db.refresh(export_record)
        return export_record

    async def list_exports(
        self,
        db: AsyncSession,
        *,
        page: int = 1,
        page_size: int = 20,
        export_type: Optional[ExportType] = None,
        export_format: Optional[ExportFormat] = None,
        status: Optional[ExportStatus] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ) -> Tuple[List[ExportHistory], int]:
        """
        List export records with pagination and filters.

        Args:
            db: Database session
            page: Page number
            page_size: Items per page
            export_type: Filter by export type
            export_format: Filter by export format
            status: Filter by status
            start_date: Filter by start date
            end_date: Filter by end date

        Returns:
            Tuple of (list of ExportHistory, total count)
        """
        filters = []

        if export_type:
            filters.append(ExportHistory.export_type == export_type)
        if export_format:
            filters.append(ExportHistory.export_format == export_format)
        if status:
            filters.append(ExportHistory.status == status)
        if start_date:
            filters.append(ExportHistory.created_at >= start_date)
        if end_date:
            filters.append(ExportHistory.created_at <= end_date)

        # Count query
        count_query = select(func.count()).select_from(ExportHistory)
        if filters:
            count_query = count_query.where(and_(*filters))
        count_result = await db.execute(count_query)
        total = count_result.scalar_one()

        # Data query
        skip = (page - 1) * page_size
        query = (
            select(ExportHistory)
            .where(and_(*filters)) if filters else select(ExportHistory)
            .order_by(ExportHistory.created_at.desc())
            .offset(skip)
            .limit(page_size)
        )
        result = await db.execute(query)
        records = list(result.scalars().all())

        return records, total

    async def export_hotels_to_excel(
        self,
        db: AsyncSession,
        *,
        hotel_ids: Optional[List[str]] = None,
        use_template: bool = False,
    ) -> Tuple[bytes, str]:
        """
        Export hotels to Excel format.

        Args:
            db: Database session
            hotel_ids: Optional list of hotel IDs to filter
            use_template: Whether to use Expedia template format

        Returns:
            Tuple of (file bytes, file name)
        """
        # Query hotels
        query = select(Hotel)
        if hotel_ids:
            query = query.where(Hotel.id.in_(hotel_ids))
        result = await db.execute(query)
        hotels = list(result.scalars().all())

        if use_template:
            # Use ExcelTemplateGenerator for template-based export
            generator = get_excel_template_generator()
            wb = generator.generate_hotel_template()
            # Fill in actual data
            ws = wb.active
            for idx, hotel in enumerate(hotels, start=2):
                ws.cell(row=idx, column=1, value=hotel.name_cn)
                ws.cell(row=idx, column=2, value=hotel.name_en)
                ws.cell(row=idx, column=3, value=hotel.brand.value if hotel.brand else None)
                ws.cell(row=idx, column=4, value=hotel.country_code)
                ws.cell(row=idx, column=5, value=hotel.province)
                ws.cell(row=idx, column=6, value=hotel.city)
                ws.cell(row=idx, column=7, value=hotel.address_cn)
                ws.cell(row=idx, column=8, value=hotel.phone)
        else:
            # Simple export
            wb = Workbook()
            ws = wb.active
            ws.title = "Hotels"

            # Headers
            headers = [
                "ID", "Name (CN)", "Name (EN)", "Brand", "Status",
                "Country", "Province", "City", "District", "Address (CN)",
                "Address (EN)", "Phone", "Email", "Expedia Hotel ID"
            ]
            ws.append(headers)

            # Data
            for hotel in hotels:
                ws.append([
                    hotel.id,
                    hotel.name_cn,
                    hotel.name_en,
                    hotel.brand.value if hotel.brand else None,
                    hotel.status.value if hotel.status else None,
                    hotel.country_code,
                    hotel.province,
                    hotel.city,
                    hotel.district,
                    hotel.address_cn,
                    hotel.address_en,
                    hotel.phone,
                    hotel.email,
                    hotel.expedia_hotel_id,
                ])

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue(), f"hotels_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    async def export_hotels_to_csv(
        self,
        db: AsyncSession,
        *,
        hotel_ids: Optional[List[str]] = None,
    ) -> Tuple[bytes, str]:
        """
        Export hotels to CSV format.

        Args:
            db: Database session
            hotel_ids: Optional list of hotel IDs to filter

        Returns:
            Tuple of (file bytes, file name)
        """
        query = select(Hotel)
        if hotel_ids:
            query = query.where(Hotel.id.in_(hotel_ids))
        result = await db.execute(query)
        hotels = list(result.scalars().all())

        output = BytesIO()
        writer = csv.writer(output)

        # Headers
        writer.writerow([
            "ID", "Name (CN)", "Name (EN)", "Brand", "Status",
            "Country", "Province", "City", "District", "Address (CN)",
            "Address (EN)", "Phone", "Email", "Expedia Hotel ID"
        ])

        # Data
        for hotel in hotels:
            writer.writerow([
                hotel.id,
                hotel.name_cn,
                hotel.name_en,
                hotel.brand.value if hotel.brand else None,
                hotel.status.value if hotel.status else None,
                hotel.country_code,
                hotel.province,
                hotel.city,
                hotel.district,
                hotel.address_cn,
                hotel.address_en,
                hotel.phone,
                hotel.email,
                hotel.expedia_hotel_id,
            ])

        output.seek(0)
        return output.getvalue(), f"hotels_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    async def export_hotels_to_json(
        self,
        db: AsyncSession,
        *,
        hotel_ids: Optional[List[str]] = None,
    ) -> Tuple[bytes, str]:
        """
        Export hotels to JSON format.

        Args:
            db: Database session
            hotel_ids: Optional list of hotel IDs to filter

        Returns:
            Tuple of (file bytes, file name)
        """
        query = select(Hotel)
        if hotel_ids:
            query = query.where(Hotel.id.in_(hotel_ids))
        result = await db.execute(query)
        hotels = list(result.scalars().all())

        data = []
        for hotel in hotels:
            data.append({
                "id": hotel.id,
                "name_cn": hotel.name_cn,
                "name_en": hotel.name_en,
                "brand": hotel.brand.value if hotel.brand else None,
                "status": hotel.status.value if hotel.status else None,
                "country_code": hotel.country_code,
                "province": hotel.province,
                "city": hotel.city,
                "district": hotel.district,
                "address_cn": hotel.address_cn,
                "address_en": hotel.address_en,
                "phone": hotel.phone,
                "email": hotel.email,
                "expedia_hotel_id": hotel.expedia_hotel_id,
                "expedia_chain_code": hotel.expedia_chain_code,
                "expedia_property_code": hotel.expedia_property_code,
                "latitude": hotel.latitude,
                "longitude": hotel.longitude,
                "created_at": hotel.created_at.isoformat() if hotel.created_at else None,
                "updated_at": hotel.updated_at.isoformat() if hotel.updated_at else None,
            })

        json_str = json.dumps(data, ensure_ascii=False, indent=2)
        return json_str.encode("utf-8"), f"hotels_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

    async def export_rooms_to_excel(
        self,
        db: AsyncSession,
        *,
        hotel_ids: Optional[List[str]] = None,
        room_ids: Optional[List[str]] = None,
        use_template: bool = False,
    ) -> Tuple[bytes, str]:
        """
        Export rooms to Excel format.

        Args:
            db: Database session
            hotel_ids: Optional list of hotel IDs to filter
            room_ids: Optional list of room IDs to filter
            use_template: Whether to use Expedia template format

        Returns:
            Tuple of (file bytes, file name)
        """
        query = select(Room).options(selectinload(Room.hotel))
        filters = []

        if hotel_ids:
            filters.append(Room.hotel_id.in_(hotel_ids))
        if room_ids:
            filters.append(Room.id.in_(room_ids))

        if filters:
            query = query.where(and_(*filters))

        result = await db.execute(query)
        rooms = list(result.scalars().all())

        if use_template:
            generator = get_excel_template_generator()
            wb = generator.generate_room_template()
            ws = wb.active
            for idx, room in enumerate(rooms, start=2):
                ws.cell(row=idx, column=1, value=room.room_type_code)
                ws.cell(row=idx, column=2, value=room.name_cn)
                ws.cell(row=idx, column=3, value=room.name_en)
                ws.cell(row=idx, column=4, value=room.bed_type)
                ws.cell(row=idx, column=5, value=room.max_occupancy)
                ws.cell(row=idx, column=6, value=room.total_rooms)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Rooms"

            headers = [
                "ID", "Hotel ID", "Hotel Name", "Room Type Code", "Name (CN)", "Name (EN)",
                "Bed Type", "Max Occupancy", "Standard Occupancy", "Room Size",
                "Floor Range", "Total Rooms", "Expedia Room ID", "Is Active"
            ]
            ws.append(headers)

            for room in rooms:
                ws.append([
                    room.id,
                    room.hotel_id,
                    room.hotel.name_cn if room.hotel else None,
                    room.room_type_code,
                    room.name_cn,
                    room.name_en,
                    room.bed_type,
                    room.max_occupancy,
                    room.standard_occupancy,
                    room.room_size,
                    room.floor_range,
                    room.total_rooms,
                    room.expedia_room_id,
                    room.is_active,
                ])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue(), f"rooms_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    async def export_rooms_to_csv(
        self,
        db: AsyncSession,
        *,
        hotel_ids: Optional[List[str]] = None,
        room_ids: Optional[List[str]] = None,
    ) -> Tuple[bytes, str]:
        """
        Export rooms to CSV format.

        Args:
            db: Database session
            hotel_ids: Optional list of hotel IDs to filter
            room_ids: Optional list of room IDs to filter

        Returns:
            Tuple of (file bytes, file name)
        """
        query = select(Room).options(selectinload(Room.hotel))
        filters = []

        if hotel_ids:
            filters.append(Room.hotel_id.in_(hotel_ids))
        if room_ids:
            filters.append(Room.id.in_(room_ids))

        if filters:
            query = query.where(and_(*filters))

        result = await db.execute(query)
        rooms = list(result.scalars().all())

        output = BytesIO()
        writer = csv.writer(output)

        writer.writerow([
            "ID", "Hotel ID", "Hotel Name", "Room Type Code", "Name (CN)", "Name (EN)",
            "Bed Type", "Max Occupancy", "Standard Occupancy", "Room Size",
            "Floor Range", "Total Rooms", "Expedia Room ID", "Is Active"
        ])

        for room in rooms:
            writer.writerow([
                room.id,
                room.hotel_id,
                room.hotel.name_cn if room.hotel else None,
                room.room_type_code,
                room.name_cn,
                room.name_en,
                room.bed_type,
                room.max_occupancy,
                room.standard_occupancy,
                room.room_size,
                room.floor_range,
                room.total_rooms,
                room.expedia_room_id,
                room.is_active,
            ])

        output.seek(0)
        return output.getvalue(), f"rooms_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    async def export_rooms_to_json(
        self,
        db: AsyncSession,
        *,
        hotel_ids: Optional[List[str]] = None,
        room_ids: Optional[List[str]] = None,
    ) -> Tuple[bytes, str]:
        """
        Export rooms to JSON format.

        Args:
            db: Database session
            hotel_ids: Optional list of hotel IDs to filter
            room_ids: Optional list of room IDs to filter

        Returns:
            Tuple of (file bytes, file name)
        """
        query = select(Room).options(selectinload(Room.hotel))
        filters = []

        if hotel_ids:
            filters.append(Room.hotel_id.in_(hotel_ids))
        if room_ids:
            filters.append(Room.id.in_(room_ids))

        if filters:
            query = query.where(and_(*filters))

        result = await db.execute(query)
        rooms = list(result.scalars().all())

        data = []
        for room in rooms:
            data.append({
                "id": room.id,
                "hotel_id": room.hotel_id,
                "hotel_name": room.hotel.name_cn if room.hotel else None,
                "room_type_code": room.room_type_code,
                "name_cn": room.name_cn,
                "name_en": room.name_en,
                "description_cn": room.description_cn,
                "description_en": room.description_en,
                "bed_type": room.bed_type,
                "max_occupancy": room.max_occupancy,
                "standard_occupancy": room.standard_occupancy,
                "room_size": room.room_size,
                "floor_range": room.floor_range,
                "total_rooms": room.total_rooms,
                "expedia_room_id": room.expedia_room_id,
                "expedia_room_type_code": room.expedia_room_type_code,
                "is_active": room.is_active,
                "created_at": room.created_at.isoformat() if room.created_at else None,
                "updated_at": room.updated_at.isoformat() if room.updated_at else None,
            })

        json_str = json.dumps(data, ensure_ascii=False, indent=2)
        return json_str.encode("utf-8"), f"rooms_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

    async def increment_download_count(
        self, db: AsyncSession, *, export_id: str
    ) -> Optional[ExportHistory]:
        """
        Increment download count for an export record.

        Args:
            db: Database session
            export_id: Export record ID

        Returns:
            Updated ExportHistory or None
        """
        export_record = await self.get_export_record(db, export_id=export_id)
        if not export_record:
            return None

        export_record.download_count += 1
        export_record.last_downloaded_at = datetime.now()

        db.add(export_record)
        await db.flush()
        await db.refresh(export_record)
        return export_record


# Singleton instance
_export_service: Optional[ExportService] = None


def get_export_service() -> ExportService:
    """Get export service singleton."""
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service
