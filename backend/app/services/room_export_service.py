"""
Room export service.

Handles room data export to Excel/CSV files with template support.
"""

import csv
import json
import os
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    openpyxl = None  # type: ignore

from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.export_history import ExportFormat, ExportHistory, ExportStatus, ExportType
from app.models.expedia_template import FieldMapping, FieldMappingType
from app.models.hotel import Hotel, Room
from app.models.room import RoomExtension
from app.services.field_mapping_service import field_mapping_service


class RoomExportService:
    """
    Service for room data export operations.

    Provides functionality to:
    - Export room data to Excel/CSV files
    - Apply field mappings for Expedia template exports
    - Track export history and statistics
    """

    # Default column order for room export
    DEFAULT_ROOM_COLUMNS = [
        ("hotel_id", "Hotel ID"),
        ("hotel_name_cn", "Hotel Name (CN)"),
        ("hotel_name_en", "Hotel Name (EN)"),
        ("room_id", "Room ID"),
        ("room_type_code", "Room Type Code"),
        ("name_cn", "Room Name (CN)"),
        ("name_en", "Room Name (EN)"),
        ("description_cn", "Description (CN)"),
        ("description_en", "Description (EN)"),
        ("bed_type", "Bed Type"),
        ("max_occupancy", "Max Occupancy"),
        ("standard_occupancy", "Standard Occupancy"),
        ("room_size", "Room Size (m2)"),
        ("floor_range", "Floor Range"),
        ("total_rooms", "Total Rooms"),
        ("expedia_room_id", "Expedia Room ID"),
        ("expedia_room_type_code", "Expedia Room Type Code"),
        ("is_active", "Is Active"),
        # Extension fields
        ("amenities_cn", "Amenities (CN)"),
        ("amenities_en", "Amenities (EN)"),
        ("view_type", "View Type"),
        ("balcony", "Balcony"),
        ("smoking_policy", "Smoking Policy"),
        ("bathroom_type", "Bathroom Type"),
        ("floor", "Floor"),
    ]

    def __init__(self):
        """Initialize RoomExportService."""
        pass

    async def _get_room_data_with_hotel(
        self,
        db: AsyncSession,
        room_ids: Optional[List[str]] = None,
        hotel_ids: Optional[List[str]] = None,
        is_active: Optional[bool] = None,
    ) -> List[Dict[str, Any]]:
        """
        Get room data with hotel information.

        Args:
            db: Database session
            room_ids: Optional list of room IDs to filter
            hotel_ids: Optional list of hotel IDs to filter
            is_active: Optional filter by active status

        Returns:
            List of room data dictionaries
        """
        # Build query
        query = select(Room).options(
            selectinload(Room.hotel),
        ).join(Hotel)

        filters = []

        if room_ids:
            filters.append(Room.id.in_(room_ids))
        if hotel_ids:
            filters.append(Room.hotel_id.in_(hotel_ids))
        if is_active is not None:
            filters.append(Room.is_active == is_active)

        if filters:
            query = query.where(and_(*filters))

        query = query.order_by(Hotel.name_cn, Room.name_cn)

        result = await db.execute(query)
        rooms = result.scalars().all()

        # Get room extensions
        room_ids_list = [r.id for r in rooms]
        extensions_query = select(RoomExtension).where(RoomExtension.room_id.in_(room_ids_list))
        extensions_result = await db.execute(extensions_query)
        extensions = {ext.room_id: ext for ext in extensions_result.scalars().all()}

        # Build result data
        room_data_list = []
        for room in rooms:
            hotel = room.hotel
            extension = extensions.get(room.id)

            data = {
                # Hotel info
                "hotel_id": hotel.id if hotel else None,
                "hotel_name_cn": hotel.name_cn if hotel else None,
                "hotel_name_en": hotel.name_en if hotel else None,
                # Room info
                "room_id": room.id,
                "room_type_code": room.room_type_code,
                "name_cn": room.name_cn,
                "name_en": room.name_en,
                "description_cn": room.description_cn,
                "description_en": room.description_en,
                "bed_type": room.bed_type,
                "max_occupancy": room.max_occupancy,
                "standard_occupancy": room.standard_occupancy,
                "room_size": room.room_size,
                "floor_range": room.floor_range,
                "total_rooms": room.total_rooms,
                "expedia_room_id": room.expedia_room_id,
                "expedia_room_type_code": room.expedia_room_type_code,
                "is_active": room.is_active,
            }

            # Add extension data if available
            if extension:
                data.update({
                    "amenities_cn": extension.amenities_cn,
                    "amenities_en": extension.amenities_en,
                    "view_type": extension.view_type,
                    "balcony": extension.balcony,
                    "smoking_policy": extension.smoking_policy,
                    "bathroom_type": extension.bathroom_type,
                    "floor": extension.floor,
                })
            else:
                data.update({
                    "amenities_cn": None,
                    "amenities_en": None,
                    "view_type": None,
                    "balcony": None,
                    "smoking_policy": None,
                    "bathroom_type": None,
                    "floor": None,
                })

            room_data_list.append(data)

        return room_data_list

    def _apply_field_mapping(
        self,
        room_data: List[Dict[str, Any]],
        mappings: List[FieldMapping],
    ) -> List[Dict[str, Any]]:
        """
        Apply field mappings to room data.

        Args:
            room_data: List of room data dictionaries
            mappings: List of field mappings

        Returns:
            List of transformed data dictionaries
        """
        if not mappings:
            return room_data

        result = []
        for row in room_data:
            new_row: Dict[str, Any] = {}

            for mapping in mappings:
                if not mapping.is_active:
                    continue

                source_field = mapping.source_field
                target_field = mapping.target_field

                if mapping.mapping_type == FieldMappingType.DIRECT:
                    new_row[target_field] = row.get(source_field, mapping.default_value)

                elif mapping.mapping_type == FieldMappingType.FIXED:
                    new_row[target_field] = mapping.default_value

                elif mapping.mapping_type == FieldMappingType.NULL:
                    new_row[target_field] = None

                elif mapping.mapping_type == FieldMappingType.LOOKUP:
                    # Lookup from mapping_config
                    if mapping.mapping_config:
                        try:
                            lookup_dict = json.loads(mapping.mapping_config)
                            source_value = row.get(source_field)
                            new_row[target_field] = lookup_dict.get(source_value, mapping.default_value)
                        except json.JSONDecodeError:
                            new_row[target_field] = mapping.default_value
                    else:
                        new_row[target_field] = mapping.default_value

                elif mapping.mapping_type == FieldMappingType.TRANSFORM:
                    # Apply transformation
                    source_value = row.get(source_field)
                    if source_value is not None and mapping.transform_script:
                        try:
                            # Simple eval for transform script
                            local_vars = {"value": source_value}
                            new_row[target_field] = eval(mapping.transform_script, {"__builtins__": {}}, local_vars)
                        except Exception:
                            new_row[target_field] = source_value
                    else:
                        new_row[target_field] = source_value if source_value is not None else mapping.default_value

                elif mapping.mapping_type == FieldMappingType.COMPUTED:
                    # Computed value using transform script
                    if mapping.transform_script:
                        try:
                            local_vars = {k: v for k, v in row.items() if v is not None}
                            new_row[target_field] = eval(mapping.transform_script, {"__builtins__": {}}, local_vars)
                        except Exception:
                            new_row[target_field] = mapping.default_value
                    else:
                        new_row[target_field] = mapping.default_value

            result.append(new_row)

        return result

    async def export_to_excel(
        self,
        db: AsyncSession,
        file_name: Optional[str] = None,
        room_ids: Optional[List[str]] = None,
        hotel_ids: Optional[List[str]] = None,
        is_active: Optional[bool] = True,
        include_headers: bool = True,
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
        operator_ip: Optional[str] = None,
    ) -> Tuple[ExportHistory, bytes]:
        """
        Export room data to Excel format.

        Args:
            db: Database session
            file_name: Optional output file name
            room_ids: Optional list of room IDs to export
            hotel_ids: Optional list of hotel IDs to export
            is_active: Optional filter by active status (default True)
            include_headers: Whether to include column headers
            operator_id: Optional operator ID for audit
            operator_name: Optional operator name for audit
            operator_ip: Optional operator IP for audit

        Returns:
            Tuple of (ExportHistory record, Excel file bytes)
        """
        if openpyxl is None:
            raise ImportError("openpyxl is required for Excel export")

        # Get room data
        room_data = await self._get_room_data_with_hotel(
            db,
            room_ids=room_ids,
            hotel_ids=hotel_ids,
            is_active=is_active,
        )

        # Generate file name
        if not file_name:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            file_name = f"room_export_{timestamp}.xlsx"

        # Create export history record
        export_history = ExportHistory(
            file_name=file_name,
            export_type=ExportType.ROOM,
            export_format=ExportFormat.EXCEL,
            status=ExportStatus.PROCESSING,
            total_rows=len(room_data),
            total_hotels=len(set(r.get("hotel_id") for r in room_data if r.get("hotel_id"))),
            total_rooms=len(room_data),
            filter_criteria=json.dumps({
                "room_ids": room_ids,
                "hotel_ids": hotel_ids,
                "is_active": is_active,
            }, ensure_ascii=False),
            hotel_ids=json.dumps(hotel_ids, ensure_ascii=False) if hotel_ids else None,
            operator_id=operator_id,
            operator_name=operator_name,
            operator_ip=operator_ip,
            started_at=datetime.utcnow(),
        )
        db.add(export_history)
        await db.flush()

        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rooms"

            # Write headers
            if include_headers:
                headers = [col[1] for col in self.DEFAULT_ROOM_COLUMNS]
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Write data
            for row_idx, row_data in enumerate(room_data, start=2 if include_headers else 1):
                for col_idx, (field, _) in enumerate(self.DEFAULT_ROOM_COLUMNS, start=1):
                    value = row_data.get(field)
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col_idx in range(1, len(self.DEFAULT_ROOM_COLUMNS) + 1):
                max_length = 0
                column = openpyxl.utils.get_column_letter(col_idx)
                for row in ws.iter_rows(min_row=1 if include_headers else 2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            file_bytes = output.getvalue()

            # Update export history
            export_history.status = ExportStatus.COMPLETED
            export_history.file_size = len(file_bytes)
            export_history.completed_at = datetime.utcnow()
            export_history.processing_time = (export_history.completed_at - export_history.started_at).total_seconds()

            await db.flush()

            return export_history, file_bytes

        except Exception as e:
            export_history.status = ExportStatus.FAILED
            export_history.error_message = str(e)
            export_history.completed_at = datetime.utcnow()
            export_history.processing_time = (export_history.completed_at - export_history.started_at).total_seconds()
            await db.flush()
            raise

    async def export_to_csv(
        self,
        db: AsyncSession,
        file_name: Optional[str] = None,
        room_ids: Optional[List[str]] = None,
        hotel_ids: Optional[List[str]] = None,
        is_active: Optional[bool] = True,
        include_headers: bool = True,
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
        operator_ip: Optional[str] = None,
    ) -> Tuple[ExportHistory, bytes]:
        """
        Export room data to CSV format.

        Args:
            db: Database session
            file_name: Optional output file name
            room_ids: Optional list of room IDs to export
            hotel_ids: Optional list of hotel IDs to export
            is_active: Optional filter by active status (default True)
            include_headers: Whether to include column headers
            operator_id: Optional operator ID for audit
            operator_name: Optional operator name for audit
            operator_ip: Optional operator IP for audit

        Returns:
            Tuple of (ExportHistory record, CSV file bytes)
        """
        # Get room data
        room_data = await self._get_room_data_with_hotel(
            db,
            room_ids=room_ids,
            hotel_ids=hotel_ids,
            is_active=is_active,
        )

        # Generate file name
        if not file_name:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            file_name = f"room_export_{timestamp}.csv"

        # Create export history record
        export_history = ExportHistory(
            file_name=file_name,
            export_type=ExportType.ROOM,
            export_format=ExportFormat.CSV,
            status=ExportStatus.PROCESSING,
            total_rows=len(room_data),
            total_hotels=len(set(r.get("hotel_id") for r in room_data if r.get("hotel_id"))),
            total_rooms=len(room_data),
            filter_criteria=json.dumps({
                "room_ids": room_ids,
                "hotel_ids": hotel_ids,
                "is_active": is_active,
            }, ensure_ascii=False),
            hotel_ids=json.dumps(hotel_ids, ensure_ascii=False) if hotel_ids else None,
            operator_id=operator_id,
            operator_name=operator_name,
            operator_ip=operator_ip,
            started_at=datetime.utcnow(),
        )
        db.add(export_history)
        await db.flush()

        try:
            # Create CSV content
            output = StringIO()
            writer = csv.writer(output)

            # Write headers
            if include_headers:
                headers = [col[1] for col in self.DEFAULT_ROOM_COLUMNS]
                writer.writerow(headers)

            # Write data
            for row_data in room_data:
                row_values = [row_data.get(field) for field, _ in self.DEFAULT_ROOM_COLUMNS]
                writer.writerow(row_values)

            # Convert to bytes
            file_bytes = output.getvalue().encode("utf-8-sig")

            # Update export history
            export_history.status = ExportStatus.COMPLETED
            export_history.file_size = len(file_bytes)
            export_history.completed_at = datetime.utcnow()
            export_history.processing_time = (export_history.completed_at - export_history.started_at).total_seconds()

            await db.flush()

            return export_history, file_bytes

        except Exception as e:
            export_history.status = ExportStatus.FAILED
            export_history.error_message = str(e)
            export_history.completed_at = datetime.utcnow()
            export_history.processing_time = (export_history.completed_at - export_history.started_at).total_seconds()
            await db.flush()
            raise

    async def export_using_template(
        self,
        db: AsyncSession,
        template_id: str,
        file_name: Optional[str] = None,
        room_ids: Optional[List[str]] = None,
        hotel_ids: Optional[List[str]] = None,
        is_active: Optional[bool] = True,
        include_headers: bool = True,
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
        operator_ip: Optional[str] = None,
    ) -> Tuple[ExportHistory, bytes]:
        """
        Export room data using an Expedia template.

        Args:
            db: Database session
            template_id: Template ID to use for export
            file_name: Optional output file name
            room_ids: Optional list of room IDs to export
            hotel_ids: Optional list of hotel IDs to export
            is_active: Optional filter by active status (default True)
            include_headers: Whether to include column headers
            operator_id: Optional operator ID for audit
            operator_name: Optional operator name for audit
            operator_ip: Optional operator IP for audit

        Returns:
            Tuple of (ExportHistory record, file bytes)
        """
        if openpyxl is None:
            raise ImportError("openpyxl is required for Excel export")

        # Get template with mappings
        from app.services.expedia_template_service import expedia_template
        template = await expedia_template.get_with_mappings(db, id=int(template_id))

        if not template:
            raise ValueError(f"Template with ID {template_id} not found")

        # Get room data
        room_data = await self._get_room_data_with_hotel(
            db,
            room_ids=room_ids,
            hotel_ids=hotel_ids,
            is_active=is_active,
        )

        # Apply field mappings
        sorted_mappings = sorted(template.field_mappings, key=lambda m: m.field_order)
        mapped_data = self._apply_field_mapping(room_data, sorted_mappings)

        # Generate file name
        if not file_name:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            file_name = f"room_template_{template.code}_{timestamp}.xlsx"

        # Create export history record
        export_history = ExportHistory(
            file_name=file_name,
            export_type=ExportType.EXPEDIA_TEMPLATE,
            export_format=ExportFormat.EXCEL,
            status=ExportStatus.PROCESSING,
            total_rows=len(mapped_data),
            total_hotels=len(set(r.get("hotel_id") for r in room_data if r.get("hotel_id"))),
            total_rooms=len(mapped_data),
            filter_criteria=json.dumps({
                "room_ids": room_ids,
                "hotel_ids": hotel_ids,
                "is_active": is_active,
            }, ensure_ascii=False),
            hotel_ids=json.dumps(hotel_ids, ensure_ascii=False) if hotel_ids else None,
            template_id=template.id,
            template_name=template.name,
            template_version=template.version,
            operator_id=operator_id,
            operator_name=operator_name,
            operator_ip=operator_ip,
            started_at=datetime.utcnow(),
        )
        db.add(export_history)
        await db.flush()

        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = template.sheet_name or "Room Data"

            # Get target fields from mappings (in order)
            target_fields = [m.target_field for m in sorted_mappings if m.is_visible]
            target_fields_with_headers = [(m.target_field, m.description or m.target_field) for m in sorted_mappings if m.is_visible]

            # Write headers
            if include_headers:
                headers = [header for _, header in target_fields_with_headers]
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=template.header_row, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Write data
            data_start = template.data_start_row
            for row_idx, row_data in enumerate(mapped_data, start=data_start):
                for col_idx, field in enumerate(target_fields, start=1):
                    value = row_data.get(field)
                    # Truncate if needed
                    mapping = next((m for m in sorted_mappings if m.target_field == field), None)
                    if mapping and mapping.target_field_max_length and value:
                        if isinstance(value, str) and len(value) > mapping.target_field_max_length:
                            value = value[:mapping.target_field_max_length]
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col_idx in range(1, len(target_fields) + 1):
                max_length = 0
                column = openpyxl.utils.get_column_letter(col_idx)
                for row in ws.iter_rows(min_row=template.header_row if include_headers else data_start, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            file_bytes = output.getvalue()

            # Update export history
            export_history.status = ExportStatus.COMPLETED
            export_history.file_size = len(file_bytes)
            export_history.completed_at = datetime.utcnow()
            export_history.processing_time = (export_history.completed_at - export_history.started_at).total_seconds()

            await db.flush()

            return export_history, file_bytes

        except Exception as e:
            export_history.status = ExportStatus.FAILED
            export_history.error_message = str(e)
            export_history.completed_at = datetime.utcnow()
            export_history.processing_time = (export_history.completed_at - export_history.started_at).total_seconds()
            await db.flush()
            raise

    async def get_export_history(
        self,
        db: AsyncSession,
        skip: int = 0,
        limit: int = 20,
        export_type: Optional[ExportType] = None,
    ) -> Tuple[List[ExportHistory], int]:
        """
        Get room export history records.

        Args:
            db: Database session
            skip: Number of records to skip
            limit: Maximum number of records to return
            export_type: Optional filter by export type

        Returns:
            Tuple of (list of ExportHistory, total count)
        """
        query = select(ExportHistory).where(
            ExportHistory.export_type.in_([ExportType.ROOM, ExportType.EXPEDIA_TEMPLATE])
        )

        if export_type:
            query = query.where(ExportHistory.export_type == export_type)

        # Count total
        count_query = select(ExportHistory).where(
            ExportHistory.export_type.in_([ExportType.ROOM, ExportType.EXPEDIA_TEMPLATE])
        )
        if export_type:
            count_query = count_query.where(ExportHistory.export_type == export_type)

        count_result = await db.execute(count_query)
        total = len(list(count_result.scalars().all()))

        # Get paginated results
        query = query.order_by(ExportHistory.created_at.desc()).offset(skip).limit(limit)
        result = await db.execute(query)
        records = list(result.scalars().all())

        return records, total


# Singleton instance
_room_export_service: Optional[RoomExportService] = None


def get_room_export_service() -> RoomExportService:
    """Get room export service singleton."""
    global _room_export_service
    if _room_export_service is None:
        _room_export_service = RoomExportService()
    return _room_export_service
