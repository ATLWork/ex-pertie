"""
Hotel data export service.
Provides export functionality for hotel and room data to various formats.
"""

import csv
import json
import time
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any, BinaryIO, Dict, List, Optional, Tuple, Union

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.export_history import (
    ExportFormat,
    ExportHistory,
    ExportStatus,
    ExportType,
)
from app.models.expedia_template import FieldMapping, FieldMappingType
from app.models.hotel import Hotel, Room
from app.models.room import RoomExtension
from app.services.field_mapping_service import FieldMappingService, field_mapping_service
from app.services.hotel_service import hotel_service
from app.services.room_service import get_room_service
from app.generators.excel_template_generator import ExcelTemplateGenerator, get_excel_template_generator


class HotelExportService:
    """
    Service for exporting hotel and room data.

    Supports:
    - Export to Excel format
    - Export to CSV format
    - Export using Expedia template configurations
    - Field mapping and transformation
    - Export history tracking
    """

    # Header style constants
    HEADER_FILL_COLOR = "CCE5FF"  # Light blue background
    REQUIRED_HEADER_FILL_COLOR = "FFCCCC"  # Light red for required fields
    HEADER_FONT_COLOR = "000000"  # Black text

    def __init__(
        self,
        field_mapping_svc: Optional[FieldMappingService] = None,
        excel_generator: Optional[ExcelTemplateGenerator] = None,
    ):
        """
        Initialize the hotel export service.

        Args:
            field_mapping_svc: Field mapping service instance
            excel_generator: Excel template generator instance
        """
        self.field_mapping_service = field_mapping_svc or field_mapping_service
        self.excel_generator = excel_generator or get_excel_template_generator()

    async def _get_field_mappings(
        self,
        db: AsyncSession,
        template_id: str,
    ) -> List[FieldMapping]:
        """
        Get active field mappings for a template.

        Args:
            db: Database session
            template_id: Template ID

        Returns:
            List of active FieldMapping instances
        """
        return await self.field_mapping_service.get_active_mappings(
            db, template_id=template_id
        )

    def _apply_field_mapping(
        self,
        hotel_data: Dict[str, Any],
        mappings: List[FieldMapping],
    ) -> Dict[str, Any]:
        """
        Apply field mappings to transform hotel data.

        Args:
            hotel_data: Source hotel/room data dictionary
            mappings: List of field mappings

        Returns:
            Transformed data dictionary
        """
        result = {}

        for mapping in mappings:
            if not mapping.is_active:
                continue

            source_value = hotel_data.get(mapping.source_field)

            # Handle different mapping types
            if mapping.mapping_type == FieldMappingType.FIXED:
                result[mapping.target_field] = mapping.default_value
            elif mapping.mapping_type == FieldMappingType.NULL:
                result[mapping.target_field] = None
            elif mapping.mapping_type == FieldMappingType.DIRECT:
                result[mapping.target_field] = source_value
            elif mapping.mapping_type == FieldMappingType.TRANSFORM:
                # Apply transform script if available
                if mapping.transform_script and source_value is not None:
                    try:
                        # Simple eval for transform scripts (in production, use safe execution)
                        result[mapping.target_field] = eval(mapping.transform_script, {"value": source_value})
                    except Exception:
                        result[mapping.target_field] = source_value
                else:
                    result[mapping.target_field] = source_value
            elif mapping.mapping_type == FieldMappingType.LOOKUP:
                # Apply lookup from mapping config
                if mapping.mapping_config and source_value is not None:
                    try:
                        lookup_dict = json.loads(mapping.mapping_config)
                        result[mapping.target_field] = lookup_dict.get(source_value, source_value)
                    except Exception:
                        result[mapping.target_field] = source_value
                else:
                    result[mapping.target_field] = source_value
            elif mapping.mapping_type == FieldMappingType.COMPUTED:
                # Computed values require transform script
                if mapping.transform_script:
                    try:
                        result[mapping.target_field] = eval(mapping.transform_script, {"data": hotel_data})
                    except Exception:
                        result[mapping.target_field] = None
                else:
                    result[mapping.target_field] = None
            else:
                result[mapping.target_field] = source_value

            # Apply max length if specified
            if mapping.target_field_max_length and result.get(mapping.target_field):
                max_len = mapping.target_field_max_length
                if isinstance(result[mapping.target_field], str) and len(result[mapping.target_field]) > max_len:
                    result[mapping.target_field] = result[mapping.target_field][:max_len]

        return result

    def _hotel_to_dict(self, hotel: Hotel, include_rooms: bool = False) -> Dict[str, Any]:
        """
        Convert Hotel model to dictionary.

        Args:
            hotel: Hotel instance
            include_rooms: Whether to include rooms data

        Returns:
            Dictionary representation of hotel
        """
        data = {
            "id": hotel.id,
            "name_cn": hotel.name_cn,
            "name_en": hotel.name_en,
            "brand": hotel.brand.value if hotel.brand else None,
            "status": hotel.status.value if hotel.status else None,
            "country_code": hotel.country_code,
            "province": hotel.province,
            "city": hotel.city,
            "district": hotel.district,
            "address_cn": hotel.address_cn,
            "address_en": hotel.address_en,
            "postal_code": hotel.postal_code,
            "phone": hotel.phone,
            "email": hotel.email,
            "website": hotel.website,
            "latitude": hotel.latitude,
            "longitude": hotel.longitude,
            "expedia_hotel_id": hotel.expedia_hotel_id,
            "expedia_chain_code": hotel.expedia_chain_code,
            "expedia_property_code": hotel.expedia_property_code,
            "opened_at": hotel.opened_at.isoformat() if hotel.opened_at else None,
            "renovated_at": hotel.renovated_at.isoformat() if hotel.renovated_at else None,
        }

        if include_rooms and hotel.rooms:
            data["rooms"] = [self._room_to_dict(room) for room in hotel.rooms]

        return data

    def _room_to_dict(self, room: Room) -> Dict[str, Any]:
        """
        Convert Room model to dictionary.

        Args:
            room: Room instance

        Returns:
            Dictionary representation of room
        """
        return {
            "id": room.id,
            "hotel_id": room.hotel_id,
            "room_type_code": room.room_type_code,
            "name_cn": room.name_cn,
            "name_en": room.name_en,
            "description_cn": room.description_cn,
            "description_en": room.description_en,
            "bed_type": room.bed_type,
            "max_occupancy": room.max_occupancy,
            "standard_occupancy": room.standard_occupancy,
            "room_size": room.room_size,
            "floor_range": room.floor_range,
            "total_rooms": room.total_rooms,
            "expedia_room_id": room.expedia_room_id,
            "expedia_room_type_code": room.expedia_room_type_code,
            "is_active": room.is_active,
        }

    def _room_extension_to_dict(self, extension: RoomExtension) -> Dict[str, Any]:
        """
        Convert RoomExtension model to dictionary.

        Args:
            extension: RoomExtension instance

        Returns:
            Dictionary representation of room extension
        """
        return {
            "amenities_cn": extension.amenities_cn,
            "amenities_en": extension.amenities_en,
            "amenity_details": extension.amenity_details,
            "image_urls": extension.image_urls,
            "thumbnail_url": extension.thumbnail_url,
            "view_type": extension.view_type,
            "balcony": extension.balcony,
            "smoking_policy": extension.smoking_policy,
            "floor": extension.floor,
            "bathroom_type": extension.bathroom_type,
            "bathroom_amenities_cn": extension.bathroom_amenities_cn,
            "bathroom_amenities_en": extension.bathroom_amenities_en,
            "accessibility_features": extension.accessibility_features,
        }

    def _create_export_workbook(
        self,
        headers: List[str],
        data_rows: List[List[Any]],
        sheet_name: str = "Export Data",
    ) -> Workbook:
        """
        Create an Excel workbook with export data.

        Args:
            headers: List of column headers
            data_rows: List of data rows
            sheet_name: Name of the worksheet

        Returns:
            Workbook object
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Style definitions
        header_fill = PatternFill(
            start_color=self.HEADER_FILL_COLOR,
            end_color=self.HEADER_FILL_COLOR,
            fill_type="solid",
        )
        header_font = Font(bold=True, color=self.HEADER_FONT_COLOR)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        data_alignment = Alignment(horizontal="left", vertical="center")

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = data_alignment
                cell.border = thin_border

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(headers[col_idx - 1])) if headers[col_idx - 1] else 10
            for row_idx in range(2, len(data_rows) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        return wb

    async def export_to_excel(
        self,
        db: AsyncSession,
        *,
        hotel_ids: Optional[List[str]] = None,
        include_rooms: bool = False,
        custom_columns: Optional[List[str]] = None,
        filter_criteria: Optional[Dict[str, Any]] = None,
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
    ) -> Tuple[BinaryIO, str, ExportHistory]:
        """
        Export hotel data to Excel format.

        Args:
            db: Database session
            hotel_ids: Optional list of specific hotel IDs to export
            include_rooms: Whether to include room data
            custom_columns: Optional list of column names to export
            filter_criteria: Optional filter criteria for hotels
            operator_id: ID of the user initiating the export
            operator_name: Name of the user initiating the export

        Returns:
            Tuple of (file buffer, filename, export history record)
        """
        start_time = time.time()
        file_name = f"hotel_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Create export history record
        export_record = ExportHistory(
            file_name=file_name,
            export_type=ExportType.ROOM if include_rooms else ExportType.HOTEL,
            export_format=ExportFormat.EXCEL,
            status=ExportStatus.PROCESSING,
            filter_criteria=json.dumps(filter_criteria) if filter_criteria else None,
            hotel_ids=json.dumps(hotel_ids) if hotel_ids else None,
            operator_id=operator_id,
            operator_name=operator_name,
            started_at=datetime.now(),
        )
        db.add(export_record)
        await db.flush()

        try:
            # Fetch hotels
            if hotel_ids:
                hotels = []
                for hotel_id in hotel_ids:
                    hotel = await hotel_service.get_hotel(db, hotel_id=hotel_id)
                    if hotel:
                        hotels.append(hotel)
            elif filter_criteria:
                hotels, _ = await hotel_service.list_hotels(db, query=None, page=1, page_size=10000)
            else:
                hotels, _ = await hotel_service.list_hotels(db, query=None, page=1, page_size=10000)

            # Build data rows
            headers = []
            data_rows = []

            if include_rooms:
                # Export hotels with rooms
                headers = [
                    "Hotel ID", "Hotel Name (CN)", "Hotel Name (EN)", "Brand", "City", "Province",
                    "Address (CN)", "Address (EN)", "Phone", "Email",
                    "Expedia Hotel ID", "Expedia Chain Code", "Expedia Property Code",
                    "Room Type Code", "Room Name (CN)", "Room Name (EN)", "Bed Type",
                    "Max Occupancy", "Room Size", "Total Rooms", "Expedia Room ID",
                ]

                for hotel in hotels:
                    hotel_dict = self._hotel_to_dict(hotel)
                    if hotel.rooms:
                        for room in hotel.rooms:
                            room_dict = self._room_to_dict(room)
                            row = [
                                hotel.id,
                                hotel.name_cn,
                                hotel.name_en,
                                hotel.brand.value if hotel.brand else None,
                                hotel.city,
                                hotel.province,
                                hotel.address_cn,
                                hotel.address_en,
                                hotel.phone,
                                hotel.email,
                                hotel.expedia_hotel_id,
                                hotel.expedia_chain_code,
                                hotel.expedia_property_code,
                                room.room_type_code,
                                room.name_cn,
                                room.name_en,
                                room.bed_type,
                                room.max_occupancy,
                                room.room_size,
                                room.total_rooms,
                                room.expedia_room_id,
                            ]
                            data_rows.append(row)
                    else:
                        # Hotel with no rooms - still add a row
                        row = [
                            hotel.id,
                            hotel.name_cn,
                            hotel.name_en,
                            hotel.brand.value if hotel.brand else None,
                            hotel.city,
                            hotel.province,
                            hotel.address_cn,
                            hotel.address_en,
                            hotel.phone,
                            hotel.email,
                            hotel.expedia_hotel_id,
                            hotel.expedia_chain_code,
                            hotel.expedia_property_code,
                            None, None, None, None, None, None, None, None,
                        ]
                        data_rows.append(row)
            else:
                # Export hotels only
                headers = [
                    "Hotel ID", "Hotel Name (CN)", "Hotel Name (EN)", "Brand", "Status",
                    "Country", "Province", "City", "District", "Address (CN)", "Address (EN)",
                    "Postal Code", "Phone", "Email", "Website",
                    "Latitude", "Longitude",
                    "Expedia Hotel ID", "Expedia Chain Code", "Expedia Property Code",
                ]

                for hotel in hotels:
                    row = [
                        hotel.id,
                        hotel.name_cn,
                        hotel.name_en,
                        hotel.brand.value if hotel.brand else None,
                        hotel.status.value if hotel.status else None,
                        hotel.country_code,
                        hotel.province,
                        hotel.city,
                        hotel.district,
                        hotel.address_cn,
                        hotel.address_en,
                        hotel.postal_code,
                        hotel.phone,
                        hotel.email,
                        hotel.website,
                        hotel.latitude,
                        hotel.longitude,
                        hotel.expedia_hotel_id,
                        hotel.expedia_chain_code,
                        hotel.expedia_property_code,
                    ]
                    data_rows.append(row)

            # Filter columns if specified
            if custom_columns:
                # This would require more complex logic to map custom column names
                pass

            # Create workbook
            wb = self._create_export_workbook(
                headers=headers,
                data_rows=data_rows,
                sheet_name="Hotels" if not include_rooms else "Hotels & Rooms",
            )

            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Update export record
            processing_time = time.time() - start_time
            export_record.status = ExportStatus.COMPLETED
            export_record.total_rows = len(data_rows)
            export_record.total_hotels = len(hotels)
            export_record.total_rooms = sum(len(h.rooms) for h in hotels) if include_rooms else 0
            export_record.processing_time = processing_time
            export_record.completed_at = datetime.now()
            export_record.file_size = buffer.getbuffer().nbytes

            await db.flush()

            return buffer, file_name, export_record

        except Exception as e:
            export_record.status = ExportStatus.FAILED
            export_record.error_message = str(e)
            export_record.completed_at = datetime.now()
            export_record.processing_time = time.time() - start_time
            await db.flush()
            raise

    async def export_to_csv(
        self,
        db: AsyncSession,
        *,
        hotel_ids: Optional[List[str]] = None,
        include_rooms: bool = False,
        filter_criteria: Optional[Dict[str, Any]] = None,
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
    ) -> Tuple[StringIO, str, ExportHistory]:
        """
        Export hotel data to CSV format.

        Args:
            db: Database session
            hotel_ids: Optional list of specific hotel IDs to export
            include_rooms: Whether to include room data
            filter_criteria: Optional filter criteria for hotels
            operator_id: ID of the user initiating the export
            operator_name: Name of the user initiating the export

        Returns:
            Tuple of (string buffer, filename, export history record)
        """
        start_time = time.time()
        file_name = f"hotel_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        # Create export history record
        export_record = ExportHistory(
            file_name=file_name,
            export_type=ExportType.ROOM if include_rooms else ExportType.HOTEL,
            export_format=ExportFormat.CSV,
            status=ExportStatus.PROCESSING,
            filter_criteria=json.dumps(filter_criteria) if filter_criteria else None,
            hotel_ids=json.dumps(hotel_ids) if hotel_ids else None,
            operator_id=operator_id,
            operator_name=operator_name,
            started_at=datetime.now(),
        )
        db.add(export_record)
        await db.flush()

        try:
            # Fetch hotels
            if hotel_ids:
                hotels = []
                for hotel_id in hotel_ids:
                    hotel = await hotel_service.get_hotel(db, hotel_id=hotel_id)
                    if hotel:
                        hotels.append(hotel)
            elif filter_criteria:
                hotels, _ = await hotel_service.list_hotels(db, query=None, page=1, page_size=10000)
            else:
                hotels, _ = await hotel_service.list_hotels(db, query=None, page=1, page_size=10000)

            # Build CSV
            output = StringIO()
            writer = csv.writer(output)

            if include_rooms:
                headers = [
                    "Hotel ID", "Hotel Name (CN)", "Hotel Name (EN)", "Brand", "City", "Province",
                    "Address (CN)", "Address (EN)", "Phone", "Email",
                    "Expedia Hotel ID", "Expedia Chain Code", "Expedia Property Code",
                    "Room Type Code", "Room Name (CN)", "Room Name (EN)", "Bed Type",
                    "Max Occupancy", "Room Size", "Total Rooms", "Expedia Room ID",
                ]
            else:
                headers = [
                    "Hotel ID", "Hotel Name (CN)", "Hotel Name (EN)", "Brand", "Status",
                    "Country", "Province", "City", "District", "Address (CN)", "Address (EN)",
                    "Postal Code", "Phone", "Email", "Website",
                    "Latitude", "Longitude",
                    "Expedia Hotel ID", "Expedia Chain Code", "Expedia Property Code",
                ]

            writer.writerow(headers)

            for hotel in hotels:
                if include_rooms and hotel.rooms:
                    for room in hotel.rooms:
                        row = [
                            hotel.id,
                            hotel.name_cn,
                            hotel.name_en,
                            hotel.brand.value if hotel.brand else None,
                            hotel.city,
                            hotel.province,
                            hotel.address_cn,
                            hotel.address_en,
                            hotel.phone,
                            hotel.email,
                            hotel.expedia_hotel_id,
                            hotel.expedia_chain_code,
                            hotel.expedia_property_code,
                            room.room_type_code,
                            room.name_cn,
                            room.name_en,
                            room.bed_type,
                            room.max_occupancy,
                            room.room_size,
                            room.total_rooms,
                            room.expedia_room_id,
                        ]
                        writer.writerow(row)
                else:
                    row = [
                        hotel.id,
                        hotel.name_cn,
                        hotel.name_en,
                        hotel.brand.value if hotel.brand else None,
                        hotel.status.value if hotel.status else None,
                        hotel.country_code,
                        hotel.province,
                        hotel.city,
                        hotel.district,
                        hotel.address_cn,
                        hotel.address_en,
                        hotel.postal_code,
                        hotel.phone,
                        hotel.email,
                        hotel.website,
                        hotel.latitude,
                        hotel.longitude,
                        hotel.expedia_hotel_id,
                        hotel.expedia_chain_code,
                        hotel.expedia_property_code,
                    ]
                    writer.writerow(row)

            # Update export record
            processing_time = time.time() - start_time
            export_record.status = ExportStatus.COMPLETED
            export_record.total_rows = len(hotels) if not include_rooms else sum(len(h.rooms) for h in hotels)
            export_record.total_hotels = len(hotels)
            export_record.total_rooms = sum(len(h.rooms) for h in hotels) if include_rooms else 0
            export_record.processing_time = processing_time
            export_record.completed_at = datetime.now()

            await db.flush()

            output.seek(0)
            return output, file_name, export_record

        except Exception as e:
            export_record.status = ExportStatus.FAILED
            export_record.error_message = str(e)
            export_record.completed_at = datetime.now()
            export_record.processing_time = time.time() - start_time
            await db.flush()
            raise

    async def export_using_template(
        self,
        db: AsyncSession,
        *,
        template_id: str,
        hotel_ids: Optional[List[str]] = None,
        filter_criteria: Optional[Dict[str, Any]] = None,
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
    ) -> Tuple[BinaryIO, str, ExportHistory]:
        """
        Export hotel data using an Expedia template configuration.

        Args:
            db: Database session
            template_id: Template ID to use for field mappings
            hotel_ids: Optional list of specific hotel IDs to export
            filter_criteria: Optional filter criteria for hotels
            operator_id: ID of the user initiating the export
            operator_name: Name of the user initiating the export

        Returns:
            Tuple of (file buffer, filename, export history record)
        """
        start_time = time.time()
        file_name = f"expedia_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Create export history record
        export_record = ExportHistory(
            file_name=file_name,
            export_type=ExportType.EXPEDIA_TEMPLATE,
            export_format=ExportFormat.EXCEL,
            status=ExportStatus.PROCESSING,
            template_id=template_id,
            filter_criteria=json.dumps(filter_criteria) if filter_criteria else None,
            hotel_ids=json.dumps(hotel_ids) if hotel_ids else None,
            operator_id=operator_id,
            operator_name=operator_name,
            started_at=datetime.now(),
        )
        db.add(export_record)
        await db.flush()

        try:
            # Get field mappings for template
            mappings = await self._get_field_mappings(db, template_id)

            if not mappings:
                raise ValueError(f"No active field mappings found for template {template_id}")

            # Sort mappings by field order
            mappings = sorted(mappings, key=lambda m: m.field_order)

            # Build headers from mappings
            headers = [m.target_field for m in mappings]

            # Fetch hotels
            if hotel_ids:
                hotels = []
                for hotel_id in hotel_ids:
                    hotel = await hotel_service.get_hotel(db, hotel_id=hotel_id)
                    if hotel:
                        hotels.append(hotel)
            elif filter_criteria:
                hotels, _ = await hotel_service.list_hotels(db, query=None, page=1, page_size=10000)
            else:
                hotels, _ = await hotel_service.list_hotels(db, query=None, page=1, page_size=10000)

            # Build data rows by applying field mappings
            data_rows = []
            for hotel in hotels:
                hotel_dict = self._hotel_to_dict(hotel)
                mapped_row = self._apply_field_mapping(hotel_dict, mappings)
                row = [mapped_row.get(h.target_field) for h in mappings]
                data_rows.append(row)

            # Create workbook using template generator
            field_mappings_dict = [
                {
                    "target_field": m.target_field,
                    "source_field": m.source_field,
                    "source_field_cn": m.source_field_cn,
                    "target_field_required": m.target_field_required,
                    "target_field_max_length": m.target_field_max_length,
                    "description": m.description,
                }
                for m in mappings
            ]

            wb = self.excel_generator.generate_export_template(
                template_name=template_id,
                field_mappings=field_mappings_dict,
                include_sample=False,
            )

            # Write data rows to worksheet
            ws = wb.active
            for row_idx, row_data in enumerate(data_rows, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Update export record
            processing_time = time.time() - start_time
            export_record.status = ExportStatus.COMPLETED
            export_record.total_rows = len(data_rows)
            export_record.total_hotels = len(hotels)
            export_record.processing_time = processing_time
            export_record.completed_at = datetime.now()
            export_record.file_size = buffer.getbuffer().nbytes

            await db.flush()

            return buffer, file_name, export_record

        except Exception as e:
            export_record.status = ExportStatus.FAILED
            export_record.error_message = str(e)
            export_record.completed_at = datetime.now()
            export_record.processing_time = time.time() - start_time
            await db.flush()
            raise

    async def get_export_history(
        self,
        db: AsyncSession,
        *,
        skip: int = 0,
        limit: int = 20,
        operator_id: Optional[str] = None,
        status: Optional[ExportStatus] = None,
    ) -> Tuple[List[ExportHistory], int]:
        """
        Get export history records.

        Args:
            db: Database session
            skip: Number of records to skip
            limit: Maximum number of records to return
            operator_id: Optional filter by operator
            status: Optional filter by status

        Returns:
            Tuple of (list of ExportHistory, total count)
        """
        query = select(ExportHistory)

        if operator_id:
            query = query.where(ExportHistory.operator_id == operator_id)
        if status:
            query = query.where(ExportHistory.status == status)

        query = query.order_by(ExportHistory.started_at.desc())
        query = query.offset(skip).limit(limit)

        result = await db.execute(query)
        records = list(result.scalars().all())

        # Count total
        count_query = select(ExportHistory)
        if operator_id:
            count_query = count_query.where(ExportHistory.operator_id == operator_id)
        if status:
            count_query = count_query.where(ExportHistory.status == status)
        count_result = await db.execute(count_query)
        total = len(list(count_result.scalars().all()))

        return records, total


# Module-level singleton instance
_hotel_export_service: Optional[HotelExportService] = None


def get_hotel_export_service() -> HotelExportService:
    """
    Get the singleton HotelExportService instance.

    Returns:
        HotelExportService instance.
    """
    global _hotel_export_service
    if _hotel_export_service is None:
        _hotel_export_service = HotelExportService()
    return _hotel_export_service
