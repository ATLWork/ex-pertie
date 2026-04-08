"""
Excel file parser for hotel and room data.

Parses .xlsx and .xls format Excel files, extracts structured data
with support for sheet selection and error reporting.
"""

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

logger = logging.getLogger(__name__)


@dataclass
class ParseError:
    """Represents a parsing error for a specific row."""

    row: int
    field: Optional[str] = None
    message: str = ""
    value: Any = None

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "row": self.row,
            "field": self.field,
            "message": self.message,
            "value": self.value,
        }


@dataclass
class ParseResult:
    """Result of a parsing operation."""

    success: bool
    data: List[Dict[str, Any]] = field(default_factory=list)
    errors: List[ParseError] = field(default_factory=list)
    total_rows: int = 0
    success_rows: int = 0
    error_rows: int = 0

    def add_error(
        self, row: int, message: str, field: Optional[str] = None, value: Any = None
    ) -> None:
        """Add a parse error."""
        self.errors.append(
            ParseError(row=row, field=field, message=message, value=value)
        )
        self.success = False
        self.error_rows += 1

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "success": self.success,
            "data": self.data,
            "errors": [e.to_dict() for e in self.errors],
            "total_rows": self.total_rows,
            "success_rows": self.success_rows,
            "error_rows": self.error_rows,
        }


class ExcelParser:
    """
    Base class for Excel file parsing.

    Provides common functionality for parsing Excel files in .xlsx and .xls formats.

    Attributes:
        file_path: Path to the Excel file
        sheet_name: Name or index of the sheet to parse (default: 0)
        header_row: Row number for headers (0-indexed, default: 0)
        data_start_row: Row number where data starts (0-indexed, default: 1)
    """

    def __init__(
        self,
        file_path: Union[str, Path],
        sheet_name: Union[str, int, None] = 0,
        header_row: int = 0,
        data_start_row: int = 1,
    ) -> None:
        """
        Initialize the Excel parser.

        Args:
            file_path: Path to the Excel file
            sheet_name: Sheet name or index (default: 0 = first sheet)
            header_row: Row number for headers (0-indexed, default: 0)
            data_start_row: Row number where data starts (0-indexed, default: 1)
        """
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.data_start_row = data_start_row
        self._workbook = None
        self._worksheet = None
        self._headers: List[str] = []

    def _get_engine(self) -> str:
        """
        Determine the appropriate Excel engine based on file extension.

        Returns:
            Engine name: 'openpyxl' for .xlsx, 'xlrd' for .xls
        """
        suffix = self.file_path.suffix.lower()
        if suffix == ".xlsx":
            return "openpyxl"
        elif suffix == ".xls":
            return "xlrd"
        else:
            raise ValueError(f"Unsupported file format: {suffix}. Use .xlsx or .xls")

    def _open_workbook(self) -> Any:
        """
        Open the Excel workbook using appropriate engine.

        Returns:
            Open workbook object
        """
        engine = self._get_engine()

        if engine == "openpyxl":
            from openpyxl import load_workbook

            return load_workbook(self.file_path, data_only=True)
        else:  # xlrd
            import xlrd

            return xlrd.open_workbook(self.file_path)

    def _get_sheet(
        self, workbook: Any, sheet_name: Union[str, int, None]
    ) -> Tuple[Any, str]:
        """
        Get the specified sheet from the workbook.

        Args:
            workbook: Open workbook object
            sheet_name: Sheet name or index

        Returns:
            Tuple of (worksheet, sheet_name)
        """
        engine = self._get_engine()

        if engine == "openpyxl":
            if sheet_name is None:
                sheet_name = workbook.sheetnames[0]
            elif isinstance(sheet_name, int):
                sheet_name = workbook.sheetnames[sheet_name]
            return workbook[sheet_name], sheet_name
        else:  # xlrd
            if sheet_name is None:
                sheet_index = 0
            elif isinstance(sheet_name, str):
                sheet_index = workbook.sheet_names().index(sheet_name)
            else:
                sheet_index = sheet_name
            sheet = workbook.sheet_by_index(sheet_index)
            return sheet, workbook.sheet_names()[sheet_index]

    def _read_headers(self, worksheet: Any) -> List[str]:
        """
        Read headers from the header row.

        Args:
            worksheet: Open worksheet object

        Returns:
            List of header strings
        """
        engine = self._get_engine()

        if engine == "openpyxl":
            row_values = list(worksheet.iter_rows(min_row=self.header_row + 1, max_row=self.header_row + 1, values_only=True))[0]
        else:  # xlrd
            row_values = worksheet.row_values(self.header_row)

        return [str(h).strip() if h is not None else f"column_{i}" for i, h in enumerate(row_values)]

    def _read_row(self, worksheet: Any, row_index: int) -> List[Any]:
        """
        Read a row from the worksheet.

        Args:
            worksheet: Open worksheet object
            row_index: Row index (0-indexed)

        Returns:
            List of cell values
        """
        engine = self._get_engine()

        if engine == "openpyxl":
            row_values = list(
                worksheet.iter_rows(
                    min_row=row_index + 1, max_row=row_index + 1, values_only=True
                )
            )[0]
            return list(row_values)
        else:  # xlrd
            return worksheet.row_values(row_index)

    def _is_row_empty(self, row_values: List[Any]) -> bool:
        """
        Check if a row is empty (all values are None or empty).

        Args:
            row_values: List of cell values

        Returns:
            True if row is empty
        """
        return all(v is None or str(v).strip() == "" for v in row_values)

    def _normalize_value(self, value: Any) -> Any:
        """
        Normalize a cell value.

        Args:
            value: Raw cell value

        Returns:
            Normalized value
        """
        if value is None:
            return None
        if isinstance(value, (int, float, bool)):
            return value
        # Convert to string and strip whitespace
        return str(value).strip()

    def parse(self) -> ParseResult:
        """
        Parse the Excel file and return structured data.

        Returns:
            ParseResult containing parsed data and any errors
        """
        result = ParseResult(success=True)

        try:
            self._workbook = self._open_workbook()
            self._worksheet, actual_sheet_name = self._get_sheet(
                self._workbook, self.sheet_name
            )
            logger.info(f"Parsing sheet: {actual_sheet_name}")

            # Read headers
            self._headers = self._read_headers(self._worksheet)
            logger.debug(f"Headers: {self._headers}")

            # Get total rows
            engine = self._get_engine()
            if engine == "openpyxl":
                total_rows = self._worksheet.max_row
            else:  # xlrd
                total_rows = self._worksheet.nrows

            result.total_rows = max(0, total_rows - self.data_start_row)

            # Read data rows
            for row_index in range(self.data_start_row, total_rows):
                try:
                    row_values = self._read_row(self._worksheet, row_index)

                    # Skip empty rows
                    if self._is_row_empty(row_values):
                        continue

                    # Convert row to dictionary
                    row_dict: Dict[str, Any] = {}
                    for col_index, header in enumerate(self._headers):
                        if col_index < len(row_values):
                            row_dict[header] = self._normalize_value(
                                row_values[col_index]
                            )
                        else:
                            row_dict[header] = None

                    # Parse and validate row
                    parsed_row = self._parse_row(row_index, row_dict)
                    if parsed_row is not None:
                        result.data.append(parsed_row)
                        result.success_rows += 1
                    else:
                        result.error_rows += 1

                except Exception as e:
                    logger.warning(f"Error parsing row {row_index + 1}: {e}")
                    result.add_error(
                        row=row_index + 1, message=f"Failed to parse row: {str(e)}"
                    )
                    result.error_rows += 1

        except Exception as e:
            logger.error(f"Failed to parse Excel file: {e}")
            result.success = False
            result.add_error(row=0, message=f"Failed to open file: {str(e)}")
        finally:
            if self._workbook is not None:
                self._close_workbook()

        logger.info(
            f"Parse complete: {result.success_rows} success, {result.error_rows} errors"
        )
        return result

    def _parse_row(self, row_index: int, row_dict: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        Parse and validate a single row. Override in subclasses.

        Args:
            row_index: Row index (0-indexed)
            row_dict: Row data as dictionary

        Returns:
            Parsed row data or None if invalid
        """
        return row_dict

    def _close_workbook(self) -> None:
        """Close the workbook."""
        if self._workbook is not None:
            engine = self._get_engine()
            if engine == "openpyxl":
                self._workbook.close()
            self._workbook = None

    @property
    def headers(self) -> List[str]:
        """Get the parsed headers."""
        return self._headers

    @property
    def sheet_names(self) -> List[str]:
        """Get all sheet names in the workbook."""
        wb = self._open_workbook()
        try:
            engine = self._get_engine()
            if engine == "openpyxl":
                return wb.sheetnames
            else:  # xlrd
                return wb.sheet_names()
        finally:
            wb.close()


class HotelExcelParser(ExcelParser):
    """
    Parser for hotel data Excel files.

    Parses hotel information from Excel files with standard column mapping
    and business logic validation.
    """

    # Standard column mappings for hotel data
    COLUMN_MAPPINGS = {
        # Basic Info
        "hotel_id": ["hotel_id", "hotelid", "id"],
        "name_cn": ["name_cn", "name_cn", "hotel_name_cn", "hotelnamecn", "name_c"],
        "name_en": ["name_en", "name_en", "hotel_name_en", "hotelnameen", "name_e"],
        "brand": ["brand", "brand_code", "hotel_brand"],
        "status": ["status", "hotel_status"],
        # Location Info
        "country_code": ["country_code", "countrycode", "country"],
        "province": ["province", "province_name"],
        "city": ["city", "city_name"],
        "district": ["district", "district_name"],
        "address_cn": ["address_cn", "address_c", "addresscn", "addr_cn"],
        "address_en": ["address_en", "address_e", "addressen", "addr_en"],
        "postal_code": ["postal_code", "postalcode", "zip", "zipcode"],
        # Contact Info
        "phone": ["phone", "tel", "telephone", "contact_phone"],
        "email": ["email", "email_address", "contact_email"],
        "website": ["website", "url", "hotel_website"],
        # Geolocation
        "latitude": ["latitude", "lat", "y"],
        "longitude": ["longitude", "lng", "lon", "x"],
        # Expedia specific
        "expedia_hotel_id": ["expedia_hotel_id", "expedia_hotelid", "expedia_id"],
        "expedia_chain_code": ["expedia_chain_code", "chain_code"],
        "expedia_property_code": ["expedia_property_code", "property_code"],
        # Timestamps
        "opened_at": ["opened_at", "open_date", "opening_date"],
        "renovated_at": ["renovated_at", "renovation_date"],
    }

    # Required fields for hotel data
    REQUIRED_FIELDS = ["name_cn", "province", "city", "address_cn"]

    # Field types for conversion
    FIELD_TYPES = {
        "latitude": float,
        "longitude": float,
        "opened_at": str,
        "renovated_at": str,
    }

    def __init__(
        self,
        file_path: Union[str, Path],
        sheet_name: Union[str, int, None] = 0,
        header_row: int = 0,
        data_start_row: int = 1,
        column_mapping: Optional[Dict[str, List[str]]] = None,
    ) -> None:
        """
        Initialize the hotel Excel parser.

        Args:
            file_path: Path to the Excel file
            sheet_name: Sheet name or index (default: 0)
            header_row: Row number for headers (0-indexed, default: 0)
            data_start_row: Row number where data starts (0-indexed, default: 1)
            column_mapping: Custom column mapping (optional)
        """
        super().__init__(file_path, sheet_name, header_row, data_start_row)
        self.column_mapping = column_mapping or self.COLUMN_MAPPINGS
        self._normalized_headers: Dict[str, str] = {}

    def _normalize_headers(self, headers: List[str]) -> Dict[str, str]:
        """
        Normalize headers by mapping them to standard field names.

        Args:
            headers: Original headers from Excel file

        Returns:
            Mapping from original header to normalized field name
        """
        normalized: Dict[str, str] = {}
        for header in headers:
            header_lower = header.lower().strip()
            for field_name, variations in self.column_mapping.items():
                if header_lower in [v.lower() for v in variations]:
                    normalized[header] = field_name
                    break
        return normalized

    def _parse_row(self, row_index: int, row_dict: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        Parse and validate a hotel data row.

        Args:
            row_index: Row index (0-indexed)
            row_dict: Row data as dictionary

        Returns:
            Parsed hotel data or None if invalid
        """
        # Normalize headers on first row
        if not self._normalized_headers:
            self._normalized_headers = self._normalize_headers(list(row_dict.keys()))

        # Rename keys to normalized field names
        parsed_row: Dict[str, Any] = {}
        for original_key, value in row_dict.items():
            normalized_key = self._normalized_headers.get(original_key, original_key)
            parsed_row[normalized_key] = value

        # Apply field type conversions
        for field_name, field_type in self.FIELD_TYPES.items():
            if field_name in parsed_row and parsed_row[field_name] is not None:
                try:
                    if field_type == float:
                        parsed_row[field_name] = float(parsed_row[field_name])
                    elif field_type == int:
                        parsed_row[field_name] = int(parsed_row[field_name])
                except (ValueError, TypeError):
                    logger.debug(f"Could not convert {field_name} to {field_type}")

        # Validate required fields
        for required_field in self.REQUIRED_FIELDS:
            value = parsed_row.get(required_field)
            if not value or (isinstance(value, str) and not value.strip()):
                logger.warning(
                    f"Row {row_index + 1}: Missing required field '{required_field}'"
                )
                # Not adding to errors here as per base class pattern
                # Subclass can handle validation separately

        return parsed_row

    def parse(self) -> ParseResult:
        """
        Parse the hotel Excel file.

        Returns:
            ParseResult containing parsed hotel data
        """
        return super().parse()


class RoomExcelParser(ExcelParser):
    """
    Parser for room data Excel files.

    Parses room/inventory information from Excel files with standard column mapping
    and business logic validation.
    """

    # Standard column mappings for room data
    COLUMN_MAPPINGS = {
        # Basic Info
        "room_id": ["room_id", "roomid", "id"],
        "hotel_id": ["hotel_id", "hotelid", "parent_hotel_id"],
        "room_type_code": ["room_type_code", "roomtypecode", "room_code", "type_code"],
        "name_cn": ["name_cn", "name_c", "room_name_cn", "roomnamecn"],
        "name_en": ["name_en", "name_e", "room_name_en", "roomnameen"],
        "description_cn": ["description_cn", "desc_cn", "descriptioncn"],
        "description_en": ["description_en", "desc_en", "descriptionen"],
        # Room Details
        "bed_type": ["bed_type", "bedtype", "bed"],
        "max_occupancy": ["max_occupancy", "maxocc", "occupancy_max"],
        "standard_occupancy": ["standard_occupancy", "stdocc", "occupancy_standard"],
        "room_size": ["room_size", "roomsize", "size"],
        "floor_range": ["floor_range", "floorrange", "floor"],
        "total_rooms": ["total_rooms", "totalrooms", "room_count", "rooms"],
        # Expedia specific
        "expedia_room_id": ["expedia_room_id", "expedia_roomid"],
        "expedia_room_type_code": ["expedia_room_type_code", "expedia_roomtypecode"],
        # Extension fields
        "amenities_cn": ["amenities_cn", "amenitiesc", "amenities_cn"],
        "amenities_en": ["amenities_en", "amenities_e", "amenities_en"],
        "view_type": ["view_type", "viewtype", "view"],
        "balcony": ["balcony", "has_balcony"],
        "smoking_policy": ["smoking_policy", "smoking", "smoke_policy"],
        "bathroom_type": ["bathroom_type", "bathroomtype", "bathroom"],
        "is_active": ["is_active", "active", "status"],
    }

    # Required fields for room data
    REQUIRED_FIELDS = ["hotel_id", "room_type_code", "name_cn"]

    # Field types for conversion
    FIELD_TYPES = {
        "room_size": float,
        "max_occupancy": int,
        "standard_occupancy": int,
        "total_rooms": int,
        "balcony": bool,
        "is_active": bool,
    }

    def __init__(
        self,
        file_path: Union[str, Path],
        sheet_name: Union[str, int, None] = 0,
        header_row: int = 0,
        data_start_row: int = 1,
        column_mapping: Optional[Dict[str, List[str]]] = None,
    ) -> None:
        """
        Initialize the room Excel parser.

        Args:
            file_path: Path to the Excel file
            sheet_name: Sheet name or index (default: 0)
            header_row: Row number for headers (0-indexed, default: 0)
            data_start_row: Row number where data starts (0-indexed, default: 1)
            column_mapping: Custom column mapping (optional)
        """
        super().__init__(file_path, sheet_name, header_row, data_start_row)
        self.column_mapping = column_mapping or self.COLUMN_MAPPINGS
        self._normalized_headers: Dict[str, str] = {}

    def _normalize_headers(self, headers: List[str]) -> Dict[str, str]:
        """
        Normalize headers by mapping them to standard field names.

        Args:
            headers: Original headers from Excel file

        Returns:
            Mapping from original header to normalized field name
        """
        normalized: Dict[str, str] = {}
        for header in headers:
            header_lower = header.lower().strip()
            for field_name, variations in self.column_mapping.items():
                if header_lower in [v.lower() for v in variations]:
                    normalized[header] = field_name
                    break
        return normalized

    def _parse_row(self, row_index: int, row_dict: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        Parse and validate a room data row.

        Args:
            row_index: Row index (0-indexed)
            row_dict: Row data as dictionary

        Returns:
            Parsed room data or None if invalid
        """
        # Normalize headers on first row
        if not self._normalized_headers:
            self._normalized_headers = self._normalize_headers(list(row_dict.keys()))

        # Rename keys to normalized field names
        parsed_row: Dict[str, Any] = {}
        for original_key, value in row_dict.items():
            normalized_key = self._normalized_headers.get(original_key, original_key)
            parsed_row[normalized_key] = value

        # Apply field type conversions
        for field_name, field_type in self.FIELD_TYPES.items():
            if field_name in parsed_row and parsed_row[field_name] is not None:
                try:
                    value = parsed_row[field_name]
                    if field_type == bool:
                        # Handle string boolean values
                        if isinstance(value, str):
                            parsed_row[field_name] = value.lower() in ("true", "yes", "1", "y")
                        else:
                            parsed_row[field_name] = bool(value)
                    elif field_type == float:
                        parsed_row[field_name] = float(value)
                    elif field_type == int:
                        parsed_row[field_name] = int(value)
                except (ValueError, TypeError):
                    logger.debug(f"Could not convert {field_name} to {field_type}")

        # Validate required fields
        for required_field in self.REQUIRED_FIELDS:
            value = parsed_row.get(required_field)
            if not value or (isinstance(value, str) and not value.strip()):
                logger.warning(
                    f"Row {row_index + 1}: Missing required field '{required_field}'"
                )

        return parsed_row

    def parse(self) -> ParseResult:
        """
        Parse the room Excel file.

        Returns:
            ParseResult containing parsed room data
        """
        return super().parse()
