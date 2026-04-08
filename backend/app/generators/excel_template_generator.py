"""
Excel template generator for hotel and room data import/export.
"""

from dataclasses import dataclass
from typing import Any, BinaryIO, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@dataclass
class ColumnDefinition:
    """Column definition for template generation."""

    field: str  # Internal field name
    header: str  # Display header in Excel
    header_cn: str  # Chinese header name
    required: bool = False  # Whether field is required
    max_length: Optional[int] = None  # Max length for validation
    default_value: Optional[Any] = None  # Default value
    validation_rule: Optional[str] = None  # Validation rule description
    example: Optional[str] = None  # Example value


class ExcelTemplateGenerator:
    """
    Generator for Excel import/export templates.

    Supports:
    - Hotel import template generation
    - Room import template generation
    - Export template generation based on Expedia template configuration
    - Custom column selection
    - Styled headers (bold, background color)
    """

    # Header style constants
    HEADER_FILL_COLOR = "CCE5FF"  # Light blue background
    REQUIRED_HEADER_FILL_COLOR = "FFCCCC"  # Light red for required fields
    HEADER_FONT_COLOR = "000000"  # Black text

    # Default column widths
    DEFAULT_COLUMN_WIDTH = 20
    HEADER_COLUMN_WIDTH = 25

    def __init__(self):
        """Initialize the Excel template generator."""
        self._header_fill = PatternFill(
            start_color=self.HEADER_FILL_COLOR,
            end_color=self.HEADER_FILL_COLOR,
            fill_type="solid",
        )
        self._required_header_fill = PatternFill(
            start_color=self.REQUIRED_HEADER_FILL_COLOR,
            end_color=self.REQUIRED_HEADER_FILL_COLOR,
            fill_type="solid",
        )
        self._header_font = Font(
            bold=True,
            color=self.HEADER_FONT_COLOR,
        )
        self._header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        self._thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def _apply_header_style(
        self,
        cell,
        is_required: bool = False,
    ) -> None:
        """Apply styling to a header cell."""
        cell.font = self._header_font
        cell.alignment = self._header_alignment
        cell.border = self._thin_border
        if is_required:
            cell.fill = self._required_header_fill
        else:
            cell.fill = self._header_fill

    def _apply_data_cell_style(self, cell) -> None:
        """Apply styling to a data cell."""
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )
        cell.border = self._thin_border

    def _set_column_widths(
        self,
        worksheet,
        columns: List[ColumnDefinition],
    ) -> None:
        """Set column widths based on header names."""
        for idx, col_def in enumerate(columns, start=1):
            col_letter = get_column_letter(idx)
            # Use longer of header or header_cn for width calculation
            header_text = col_def.header_cn if len(col_def.header_cn) > len(col_def.header) else col_def.header
            width = max(len(header_text) + 2, self.DEFAULT_COLUMN_WIDTH)
            worksheet.column_dimensions[col_letter].width = width

    def _add_sample_rows(
        self,
        worksheet,
        columns: List[ColumnDefinition],
        start_row: int = 2,
        num_samples: int = 3,
    ) -> None:
        """Add sample data rows for user reference."""
        sample_data = [
            ["示例酒店 1", "Sample Hotel 1", "ATOUR", "CN", "上海市", "静安区", "南京西路168号", "+86-21-12345678", "hotel1@example.com", "31.2304", "121.4737", "ATOUR", "HP12345"],
            ["示例酒店 2", "Sample Hotel 2", "ATOURX", "CN", "北京市", "朝阳区", "建国路88号", "+86-10-12345678", "hotel2@example.com", "39.9042", "116.4074", "ATOURX", "HP12346"],
            ["示例酒店 3", "Sample Hotel 3", "ZHOTEL", "CN", "杭州市", "西湖区", "西湖大道1号", "+86-571-12345678", "hotel3@example.com", "30.2741", "120.1551", "ZHOTEL", "HP12347"],
        ]

        for row_idx, sample_row in enumerate(sample_data[:num_samples], start=start_row):
            for col_idx, value in enumerate(sample_row, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                self._apply_data_cell_style(cell)

    def _create_workbook(self) -> Workbook:
        """Create a new workbook with a single sheet."""
        wb = Workbook()
        ws = wb.active
        return wb, ws

    def generate_hotel_template(
        self,
        custom_columns: Optional[List[str]] = None,
        include_sample: bool = True,
        sheet_name: str = "酒店信息",
    ) -> Workbook:
        """
        Generate hotel import template.

        Args:
            custom_columns: List of field names to include. If None, all columns are included.
            include_sample: Whether to include sample data rows.
            sheet_name: Name of the worksheet.

        Returns:
            Workbook object ready to be saved.
        """
        # Define all hotel columns
        all_columns: List[ColumnDefinition] = [
            ColumnDefinition(
                field="name_cn",
                header="Hotel Name (CN)",
                header_cn="酒店名称（中文）",
                required=True,
                max_length=255,
                example="亚朵酒店",
            ),
            ColumnDefinition(
                field="name_en",
                header="Hotel Name (EN)",
                header_cn="酒店名称（英文）",
                required=False,
                max_length=255,
                example="Atour Hotel",
            ),
            ColumnDefinition(
                field="brand",
                header="Brand",
                header_cn="品牌",
                required=True,
                validation_rule="atour/atour_x/zhotel/ahaus",
                example="atour",
            ),
            ColumnDefinition(
                field="country_code",
                header="Country Code",
                header_cn="国家代码",
                required=True,
                max_length=10,
                default_value="CN",
                example="CN",
            ),
            ColumnDefinition(
                field="province",
                header="Province",
                header_cn="省份",
                required=True,
                max_length=100,
                example="上海市",
            ),
            ColumnDefinition(
                field="city",
                header="City",
                header_cn="城市",
                required=True,
                max_length=100,
                example="上海市",
            ),
            ColumnDefinition(
                field="district",
                header="District",
                header_cn="区县",
                required=False,
                max_length=100,
                example="静安区",
            ),
            ColumnDefinition(
                field="address_cn",
                header="Address (CN)",
                header_cn="地址（中文）",
                required=True,
                max_length=500,
                example="南京西路168号",
            ),
            ColumnDefinition(
                field="address_en",
                header="Address (EN)",
                header_cn="地址（英文）",
                required=False,
                max_length=500,
                example="168 Nanjing West Road",
            ),
            ColumnDefinition(
                field="postal_code",
                header="Postal Code",
                header_cn="邮政编码",
                required=False,
                max_length=20,
                example="200041",
            ),
            ColumnDefinition(
                field="phone",
                header="Phone",
                header_cn="电话",
                required=False,
                max_length=50,
                example="+86-21-12345678",
            ),
            ColumnDefinition(
                field="email",
                header="Email",
                header_cn="邮箱",
                required=False,
                max_length=255,
                example="info@hotel.com",
            ),
            ColumnDefinition(
                field="website",
                header="Website",
                header_cn="网站",
                required=False,
                max_length=500,
                example="https://www.atour.com",
            ),
            ColumnDefinition(
                field="latitude",
                header="Latitude",
                header_cn="纬度",
                required=False,
                example="31.2304",
            ),
            ColumnDefinition(
                field="longitude",
                header="Longitude",
                header_cn="经度",
                required=False,
                example="121.4737",
            ),
            ColumnDefinition(
                field="expedia_chain_code",
                header="Expedia Chain Code",
                header_cn="Expedia 连锁代码",
                required=False,
                max_length=50,
                example="ATOUR",
            ),
            ColumnDefinition(
                field="expedia_property_code",
                header="Expedia Property Code",
                header_cn="Expedia 物业代码",
                required=False,
                max_length=50,
                example="HP12345",
            ),
        ]

        # Filter columns if custom_columns specified
        if custom_columns:
            columns = [col for col in all_columns if col.field in custom_columns]
        else:
            columns = all_columns

        # Create workbook
        wb, ws = self._create_workbook()
        ws.title = sheet_name

        # Write headers
        for idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=idx)
            # Use bilingual header
            cell.value = f"{col_def.header_cn}\n{col_def.header}"
            self._apply_header_style(cell, is_required=col_def.required)

            # Add validation rule as comment if available
            if col_def.validation_rule:
                cell.comment = None  # openpyxl comment support if needed

        # Add sample data
        if include_sample:
            sample_hotels = [
                ["亚朵酒店（南京西路店）", "Atour Hotel (Nanjing West Road)", "atour", "CN", "上海市", "静安区", "南京西路168号", "+86-21-12345678", "hotel1@example.com", "31.2304", "121.4737", "ATOUR", "HP12345"],
                ["亚朵X（北京国贸店）", "ATour X (Beijing Guomao)", "atour_x", "CN", "北京市", "朝阳区", "建国路88号", "+86-10-88888888", "hotel2@example.com", "39.9042", "116.4074", "ATOURX", "HP12346"],
            ]
            for row_idx, sample_row in enumerate(sample_hotels, start=2):
                for col_idx, value in enumerate(sample_row, start=1):
                    if col_idx <= len(columns):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        self._apply_data_cell_style(cell)

        # Set column widths
        self._set_column_widths(ws, columns)

        # Freeze header row
        ws.freeze_panes = "A2"

        return wb

    def generate_room_template(
        self,
        custom_columns: Optional[List[str]] = None,
        include_sample: bool = True,
        sheet_name: str = "客房信息",
    ) -> Workbook:
        """
        Generate room import template.

        Args:
            custom_columns: List of field names to include. If None, all columns are included.
            include_sample: Whether to include sample data rows.
            sheet_name: Name of the worksheet.

        Returns:
            Workbook object ready to be saved.
        """
        # Define all room columns
        all_columns: List[ColumnDefinition] = [
            ColumnDefinition(
                field="room_type_code",
                header="Room Type Code",
                header_cn="房型代码",
                required=True,
                max_length=100,
                example="STD-001",
            ),
            ColumnDefinition(
                field="name_cn",
                header="Room Name (CN)",
                header_cn="房型名称（中文）",
                required=True,
                max_length=255,
                example="高级大床房",
            ),
            ColumnDefinition(
                field="name_en",
                header="Room Name (EN)",
                header_cn="房型名称（英文）",
                required=False,
                max_length=255,
                example="Superior King Room",
            ),
            ColumnDefinition(
                field="description_cn",
                header="Description (CN)",
                header_cn="房型描述（中文）",
                required=False,
                example="房间面积约35平方米，配备1张1.8米大床",
            ),
            ColumnDefinition(
                field="description_en",
                header="Description (EN)",
                header_cn="房型描述（英文）",
                required=False,
                example="Room size approximately 35 sqm with 1 King bed",
            ),
            ColumnDefinition(
                field="bed_type",
                header="Bed Type",
                header_cn="床型",
                required=False,
                max_length=100,
                validation_rule="King/Twin/Queen/Double",
                example="King",
            ),
            ColumnDefinition(
                field="max_occupancy",
                header="Max Occupancy",
                header_cn="最大入住人数",
                required=True,
                validation_rule="1-10",
                example="2",
            ),
            ColumnDefinition(
                field="standard_occupancy",
                header="Standard Occupancy",
                header_cn="标准入住人数",
                required=True,
                validation_rule="1-10",
                example="2",
            ),
            ColumnDefinition(
                field="room_size",
                header="Room Size (sqm)",
                header_cn="房间面积（平方米）",
                required=False,
                example="35",
            ),
            ColumnDefinition(
                field="floor_range",
                header="Floor Range",
                header_cn="楼层范围",
                required=False,
                max_length=50,
                example="3-5",
            ),
            ColumnDefinition(
                field="total_rooms",
                header="Total Rooms",
                header_cn="房间总数",
                required=True,
                validation_rule=">0",
                example="20",
            ),
            ColumnDefinition(
                field="expedia_room_type_code",
                header="Expedia Room Type Code",
                header_cn="Expedia 房型代码",
                required=False,
                max_length=50,
                example="RT001",
            ),
            # Room extension fields
            ColumnDefinition(
                field="amenities_cn",
                header="Amenities (CN)",
                header_cn="设施（中文）",
                required=False,
                example="免费WiFi, 空调, 电视, 冰箱",
            ),
            ColumnDefinition(
                field="amenities_en",
                header="Amenities (EN)",
                header_cn="设施（英文）",
                required=False,
                example="Free WiFi, Air conditioning, TV, Refrigerator",
            ),
            ColumnDefinition(
                field="view_type",
                header="View Type",
                header_cn="景观",
                required=False,
                validation_rule="City/Sea/Garden/Mountain",
                example="City",
            ),
            ColumnDefinition(
                field="balcony",
                header="Has Balcony",
                header_cn="有阳台",
                required=False,
                validation_rule="Yes/No",
                example="No",
            ),
            ColumnDefinition(
                field="smoking_policy",
                header="Smoking Policy",
                header_cn="吸烟政策",
                required=False,
                validation_rule="Smoking/Non-smoking",
                example="Non-smoking",
            ),
            ColumnDefinition(
                field="bathroom_type",
                header="Bathroom Type",
                header_cn="浴室类型",
                required=False,
                validation_rule="Shared/Private/Ensuite",
                example="Ensuite",
            ),
        ]

        # Filter columns if custom_columns specified
        if custom_columns:
            columns = [col for col in all_columns if col.field in custom_columns]
        else:
            columns = all_columns

        # Create workbook
        wb, ws = self._create_workbook()
        ws.title = sheet_name

        # Write headers
        for idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.value = f"{col_def.header_cn}\n{col_def.header}"
            self._apply_header_style(cell, is_required=col_def.required)

        # Add sample data
        if include_sample:
            sample_rooms = [
                ["STD-001", "高级大床房", "Superior King Room", "房间面积约35平方米，配备1张1.8米大床", "Room size approximately 35 sqm with 1 King bed", "King", "2", "2", "35", "3-5", "20", "RT001", "免费WiFi, 空调, 电视", "Free WiFi, AC, TV", "City", "No", "Non-smoking", "Ensuite"],
                ["STD-002", "豪华双床房", "Deluxe Twin Room", "房间面积约40平方米，配备2张1.2米单人床", "Room size approximately 40 sqm with 2 Twin beds", "Twin", "2", "2", "40", "6-10", "15", "RT002", "免费WiFi, 空调, 电视, 浴缸", "Free WiFi, AC, TV, Bathtub", "City", "No", "Non-smoking", "Ensuite"],
            ]
            for row_idx, sample_row in enumerate(sample_rooms, start=2):
                for col_idx, value in enumerate(sample_row, start=1):
                    if col_idx <= len(columns):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        self._apply_data_cell_style(cell)

        # Set column widths
        self._set_column_widths(ws, columns)

        # Freeze header row
        ws.freeze_panes = "A2"

        return wb

    def generate_export_template(
        self,
        template_name: str,
        field_mappings: List[Dict[str, Any]],
        include_sample: bool = True,
        sheet_name: Optional[str] = None,
    ) -> Workbook:
        """
        Generate export template based on Expedia template field mappings.

        Args:
            template_name: Name of the template (used for sheet name if not provided).
            field_mappings: List of field mapping dictionaries with keys:
                - target_field: Expedia field name
                - target_field_required: Whether field is required
                - source_field: Internal source field name
                - source_field_cn: Chinese name of source field
                - description: Field description
            include_sample: Whether to include sample data rows.
            sheet_name: Name of the worksheet (defaults to template_name).

        Returns:
            Workbook object ready to be saved.
        """
        # Convert field mappings to column definitions
        columns: List[ColumnDefinition] = []
        for mapping in field_mappings:
            col_def = ColumnDefinition(
                field=mapping.get("source_field", ""),
                header=mapping.get("target_field", ""),
                header_cn=mapping.get("source_field_cn", mapping.get("target_field", "")),
                required=mapping.get("target_field_required", False),
                max_length=mapping.get("target_field_max_length"),
                description=mapping.get("description"),
                example=mapping.get("example"),
            )
            columns.append(col_def)

        # Create workbook
        wb, ws = self._create_workbook()
        ws.title = sheet_name or template_name

        # Write headers
        for idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=idx)
            # Show Expedia field name and description
            header_text = col_def.header
            if col_def.header_cn and col_def.header_cn != col_def.header:
                header_text = f"{col_def.header_cn}\n{col_def.header}"
            if col_def.description:
                header_text = f"{header_text}\n({col_def.description})"
            cell.value = header_text
            self._apply_header_style(cell, is_required=col_def.required)

        # Add sample data if provided
        if include_sample:
            # Generate sample rows based on examples
            sample_row = [col.example for col in columns]
            for col_idx, value in enumerate(sample_row, start=1):
                if value is not None:
                    cell = ws.cell(row=2, column=col_idx)
                    cell.value = value
                    self._apply_data_cell_style(cell)

        # Set column widths
        self._set_column_widths(ws, columns)

        # Freeze header row
        ws.freeze_panes = "A2"

        return wb

    def save_to_file(
        self,
        workbook: Workbook,
        file_path: str,
    ) -> None:
        """
        Save workbook to file.

        Args:
            workbook: Workbook object to save.
            file_path: Path to save the file.
        """
        workbook.save(file_path)

    def save_to_buffer(
        self,
        workbook: Workbook,
    ) -> BinaryIO:
        """
        Save workbook to binary buffer.

        Args:
            workbook: Workbook object to save.

        Returns:
            Binary buffer containing the Excel file.
        """
        from io import BytesIO
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer


# Module-level singleton instance
_generator_instance: Optional[ExcelTemplateGenerator] = None


def get_excel_template_generator() -> ExcelTemplateGenerator:
    """
    Get the singleton ExcelTemplateGenerator instance.

    Returns:
        ExcelTemplateGenerator instance.
    """
    global _generator_instance
    if _generator_instance is None:
        _generator_instance = ExcelTemplateGenerator()
    return _generator_instance
