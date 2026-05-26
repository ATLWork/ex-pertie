"""
Excel 导出器 - 将爬取的数据导出为 Excel 文件
"""

import os
from datetime import datetime
from typing import List, Dict, Any
from loguru import logger

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    logger.warning("openpyxl not installed. Using xlwt as fallback.")
    openpyxl = None


class HotelExporter:
    """酒店数据导出器"""

    # Excel 列定义
    COLUMNS = [
        ("name_cn", "酒店中文名", 30),
        ("name_en", "酒店英文名", 35),
        ("city", "城市", 15),
        ("province", "省份", 15),
        ("country_code", "国家代码", 12),
        ("address", "地址", 40),
        ("phone", "电话", 20),
        ("email", "邮箱", 30),
        ("latitude", "纬度", 12),
        ("longitude", "经度", 12),
        ("star_rating", "星级", 10),
        ("check_in_time", "入住时间", 12),
        ("check_out_time", "退房时间", 12),
        ("room_count", "房间数", 10),
        ("facilities", "设施服务", 50),
        ("booking_url", "Booking URL", 60),
        ("rating", "评分", 10),
    ]

    def __init__(self, output_dir: str = "output"):
        """
        初始化导出器

        Args:
            output_dir: 输出目录
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def export_to_excel(
        self,
        hotels: List[Dict[str, Any]],
        filename: str = None,
    ) -> str:
        """
        导出为 Excel 文件

        Args:
            hotels: 酒店数据列表
            filename: 文件名（不含扩展名）

        Returns:
            生成的 Excel 文件路径
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"atour_hotels_{timestamp}"

        filepath = os.path.join(self.output_dir, f"{filename}.xlsx")

        if openpyxl:
            return self._export_openpyxl(hotels, filepath)
        else:
            return self._export_xlwt(hotels, filepath)

    def _export_openpyxl(
        self,
        hotels: List[Dict[str, Any]],
        filepath: str,
    ) -> str:
        """使用 openpyxl 导出"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "亚朵酒店数据"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头
        for col_idx, (key, label, width) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # 写入数据
        for row_idx, hotel in enumerate(hotels, 2):
            for col_idx, (key, _, _) in enumerate(self.COLUMNS, 1):
                value = hotel.get(key, "")
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 自动筛选
        ws.auto_filter.ref = ws.dimensions

        wb.save(filepath)
        logger.info(f"已导出 {len(hotels)} 条数据到: {filepath}")
        return filepath

    def _export_xlwt(
        self,
        hotels: List[Dict[str, Any]],
        filepath: str,
    ) -> str:
        """使用 xlwt 导出（备用方案）"""
        try:
            import xlwt
        except ImportError:
            raise RuntimeError("请安装 openpyxl 或 xlwt: pip install openpyxl")

        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet("亚朵酒店数据")

        # 样式
        header_style = xlwt.easyxf("font: bold on; pattern: pattern solid, fore_color 0x4472C4; font: color 0xFFFFFF")
        date_style = xlwt.XFStyle()
        date_style.num_format_str = "YYYY-MM-DD HH:mm"

        # 写入表头
        for col_idx, (key, label, width) in enumerate(self.COLUMNS):
            ws.write(0, col_idx, label, header_style)

        # 写入数据
        for row_idx, hotel in enumerate(hotels, 1):
            for col_idx, (key, _, _) in enumerate(self.COLUMNS):
                value = hotel.get(key, "")
                ws.write(row_idx, col_idx, value)

        wb.save(filepath)
        logger.info(f"已导出 {len(hotels)} 条数据到: {filepath}")
        return filepath

    def export_to_csv(
        self,
        hotels: List[Dict[str, Any]],
        filename: str = None,
    ) -> str:
        """
        导出为 CSV 文件

        Args:
            hotels: 酒店数据列表
            filename: 文件名（不含扩展名）

        Returns:
            生成的 CSV 文件路径
        """
        import csv

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"atour_hotels_{timestamp}"

        filepath = os.path.join(self.output_dir, f"{filename}.csv")

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=[k for k, _, _ in self.COLUMNS])
            writer.writeheader()
            writer.writerows(hotels)

        logger.info(f"已导出 {len(hotels)} 条数据到: {filepath}")
        return filepath