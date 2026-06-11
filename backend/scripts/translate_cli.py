"""酒店主数据翻译 CLI 工具

Usage:
  python -m scripts.translate_cli --help
  python -m scripts.translate_cli by-id <hotel-uuid>
  python -m scripts.translate_cli by-search <keyword>
  python -m scripts.translate_cli by-brand atour
  python -m scripts.translate_cli by-filter --city 上海 --brand atour
  python -m scripts.translate_cli all-untranslated
"""
import asyncio
import csv
import sys
from pathlib import Path
from typing import Dict, List, Optional

import typer
from rich.console import Console
from rich.progress import Progress
from rich.table import Table
from sqlalchemy import select
from sqlalchemy.orm import selectinload

# Ensure backend is in path
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.core.database import get_db_context
from app.models.hotel import Hotel, HotelBrand, HotelStatus, Room
from app.models.room import RoomExtension
from app.models.translation import ReviewStatus, TranslationHistory, TranslationType
from app.services.translation.batch_translator import BatchHotelTranslator

console = Console()
app = typer.Typer(
    name="translate",
    help="酒店主数据翻译工具 - Translate hotel master data from Chinese to English",
)

# ---------------------------------------------------------------------------
# 通用选项（复用于所有命令）
# ---------------------------------------------------------------------------
DRY_RUN_OPTION = typer.Option(False, "--dry-run", help="Preview only, no DB write")
NO_AI_OPTION = typer.Option(
    False, "--no-ai", help="Disable AI enhancement (use machine translation only)"
)
CONCURRENCY_OPTION = typer.Option(5, "--concurrency", help="Concurrent translation workers")
EXPORT_CSV_OPTION = typer.Option(None, "--export-csv", help="Export results to CSV file")
EXPORT_EXCEL_OPTION = typer.Option(None, "--export-excel", help="Export results to Excel file")


# ---------------------------------------------------------------------------
# 辅助函数
# ---------------------------------------------------------------------------
def _find_hotel_by_id(hotels: List[Hotel], hotel_id: str) -> Optional[Hotel]:
    """在酒店列表中按 ID 查找。"""
    for h in hotels:
        if str(h.id) == hotel_id:
            return h
    return None


def get_original_text(hotel: Hotel, field_key: str) -> str:
    """从酒店对象获取字段的原文（中文）。"""
    if ":" in field_key:
        room_id, field = field_key.split(":", 1)
        room = next((r for r in hotel.rooms if str(r.id) == room_id), None)
        if not room:
            return ""
        if field in ("name_en", "description_en"):
            cn_field = field.replace("_en", "_cn")
            return getattr(room, cn_field, "") or ""
        elif field in ("amenities_en", "bathroom_amenities_en"):
            # RoomExtension 未预加载到 hotel.rooms 上，原文从酒店预加载数据中无法获取
            return ""
        return ""
    else:
        # Hotel 级别字段
        cn_field = field_key.replace("_en", "_cn") if field_key.endswith("_en") else field_key
        return str(getattr(hotel, cn_field, "") or "")
    return ""


def _truncate(text: str, max_len: int = 50) -> str:
    """截断长文本用于表格显示。"""
    if len(text) > max_len:
        return text[: max_len - 3] + "..."
    return text


def update_hotel_fields(hotel: Hotel, fields: Dict[str, dict], ext_map: Dict[str, "RoomExtension"]):
    """将翻译结果写回酒店对象（包括 Room 和 RoomExtension）。"""
    for field_key, field_info in fields.items():
        translated_value = field_info["translated"]
        if ":" in field_key:
            room_id, field = field_key.split(":", 1)
            room = next((r for r in hotel.rooms if str(r.id) == room_id), None)
            if room and field in ("name_en", "description_en"):
                setattr(room, field, translated_value)
            elif field in ("amenities_en", "bathroom_amenities_en"):
                ext = ext_map.get(room_id)
                if ext:
                    setattr(ext, field, translated_value)
        else:
            if hasattr(hotel, field_key):
                setattr(hotel, field_key, translated_value)


def export_results_to_csv(
    results: List[dict], hotels: List[Hotel], output_path: Path
):
    """导出翻译结果到 CSV 文件。"""
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["Hotel ID", "Hotel Name", "Level", "Field", "Original", "Translated", "Source"])
        for result in results:
            hotel = _find_hotel_by_id(hotels, result["hotel_id"])
            if not hotel:
                continue
            for field_key, field_info in result["fields"].items():
                original = get_original_text(hotel, field_key)
                level = field_info["level"]
                source = field_info["source"]
                translated_value = field_info["translated"]
                writer.writerow([str(hotel.id), hotel.name_cn, level, field_key, original, translated_value, source])

    console.print(f"[green]✓ CSV exported to {output_path}[/green]")


def export_results_to_excel(
    results: List[dict], hotels: List[Hotel], output_path: Path
):
    """导出翻译结果到 Excel 文件。"""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        console.print(
            "[red]Error: openpyxl is required for Excel export. "
            "Install with: pip install openpyxl[/red]"
        )
        raise typer.Exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Translations"

    # 表头
    headers = ["Hotel ID", "Hotel Name", "Level", "Field", "Original", "Translated", "Source"]
    ws.append(headers)

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for result in results:
        hotel = _find_hotel_by_id(hotels, result["hotel_id"])
        if not hotel:
            continue
        for field_key, field_info in result["fields"].items():
            original = get_original_text(hotel, field_key)
            level = field_info["level"]
            source = field_info["source"]
            translated_value = field_info["translated"]
            ws.append([str(hotel.id), hotel.name_cn, level, field_key, original, translated_value, source])

    # 自动列宽
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 80)

    wb.save(output_path)
    console.print(f"[green]✓ Excel exported to {output_path}[/green]")


# ---------------------------------------------------------------------------
# 通用翻译工作流
# ---------------------------------------------------------------------------
async def _translate_workflow(
    *,
    hotel_ids: List[str],
    hotels_for_display: List[Hotel],
    dry_run: bool,
    no_ai: bool,
    concurrency: int,
    export_csv: Optional[Path],
    export_excel: Optional[Path],
):
    """
    通用翻译工作流。

    1. 初始化 Redis（可选，失败不影响）
    2. 调用 BatchHotelTranslator 翻译
    3. Rich Table 显示预览
    4. 导出 CSV/Excel（如果指定）
    5. 确认后写回数据库 + 创建 TranslationHistory
    """
    # ---- 1. 初始化 Redis（可选） ----
    try:
        from app.core.redis import RedisService

        await RedisService.init()
        RedisService.get_client()
    except Exception:
        pass

    # ---- 2. 翻译 ----
    async with get_db_context() as db:
        translator = BatchHotelTranslator()

        with Progress(transient=True) as progress:
            task_id = progress.add_task(
                "[cyan]Translating...", total=len(hotel_ids)
            )

            def on_progress(done: int, total: int) -> None:
                progress.update(task_id, completed=done)

            results = await translator.translate_batch(
                hotel_ids, db, concurrency=concurrency,
                progress_callback=on_progress,
            )

        # ---- 3. Rich Table 预览 ----
        table = Table(title="Translation Preview", title_style="bold cyan", highlight=True)
        table.add_column("#", style="dim", width=4, justify="right")
        table.add_column("Hotel", style="cyan", no_wrap=True)
        table.add_column("Field", style="green")
        table.add_column("Original (CN)", style="yellow", max_width=50)
        table.add_column("Translated (EN)", style="magenta", max_width=50)
        table.add_column("Source", style="dim cyan", width=12)
        table.add_column("Status", style="bold", width=8)

        total_fields = 0
        total_errors = 0

        for i, result in enumerate(results, 1):
            hotel = _find_hotel_by_id(hotels_for_display, result["hotel_id"])
            hotel_name = hotel.name_cn if hotel else "Unknown"

            # 显示错误
            if result.get("errors"):
                for err in result["errors"]:
                    table.add_row(
                        str(i), hotel_name, "ERROR", "", "", "", f"[red]✗ {err}[/red]"
                    )
                    total_errors += 1

            # 显示翻译结果
            field_items = list(result["fields"].items())
            for j, (field_key, field_info) in enumerate(field_items):
                original = get_original_text(hotel, field_key) if hotel else ""
                translated_value = field_info["translated"]
                source_value = field_info["source"]

                # 按 source 着色
                if source_value == "CACHE":
                    source_display = "[dim]CACHE[/dim]"
                elif source_value == "MACHINE":
                    source_display = "[blue]MACHINE[/blue]"
                elif source_value == "AI_ENHANCED":
                    source_display = "[green]AI_ENHANCED[/green]"
                else:
                    source_display = source_value

                table.add_row(
                    str(i) if j == 0 else "",
                    hotel_name if j == 0 else "",
                    field_key,
                    _truncate(original),
                    _truncate(translated_value),
                    source_display,
                    "[green]✓[/green]",
                )
                total_fields += 1

        console.print(table)

        # 摘要
        summary_parts = [f"{len(results)} hotels, {total_fields} fields translated"]
        if total_errors:
            summary_parts.append(f"[red]{total_errors} errors[/red]")
        console.print(f"\n[bold]Summary:[/bold] {', '.join(summary_parts)}")

        # ---- 4. 导出 ----
        if export_csv:
            export_results_to_csv(results, hotels_for_display, export_csv)
        if export_excel:
            export_results_to_excel(results, hotels_for_display, export_excel)

        # ---- 5. 确认并写入数据库 ----
        if not dry_run:
            if total_errors > 0:
                console.print(
                    "[yellow]⚠ There are translation errors. Review carefully before applying.[/yellow]"
                )

            if not typer.confirm("\nApply translations to database?"):
                console.print("[yellow]Cancelled. No changes were made.[/yellow]")
                return

            # 在当前会话中重新加载酒店（含 rooms），以便写入
            stmt = (
                select(Hotel)
                .options(selectinload(Hotel.rooms))
                .where(Hotel.id.in_(hotel_ids))
            )
            r = await db.execute(stmt)
            hotels_in_session = list(r.scalars().all())
            hotel_map: Dict[str, Hotel] = {str(h.id): h for h in hotels_in_session}

            # 加载 RoomExtension（用于 amenities_en / bathroom_amenities_en 写入）
            all_room_ids: List[str] = []
            for h in hotels_in_session:
                for room in h.rooms:
                    all_room_ids.append(room.id)

            ext_map: Dict[str, RoomExtension] = {}
            if all_room_ids:
                ext_stmt = select(RoomExtension).where(
                    RoomExtension.room_id.in_(all_room_ids)
                )
                ext_result = await db.execute(ext_stmt)
                for ext in ext_result.scalars().all():
                    ext_map[ext.room_id] = ext

            # 逐个酒店写入（per-hotel commit，按计划要求）
            translation_type = TranslationType.MACHINE if no_ai else TranslationType.HYBRID
            succeeded_hotels: List[str] = []
            failed_hotels: List[tuple[str, str]] = []

            for result in results:
                hotel = hotel_map.get(result["hotel_id"])
                if not hotel:
                    failed_hotels.append((result["hotel_id"], "Hotel not reloaded in session"))
                    continue

                try:
                    update_hotel_fields(hotel, result["fields"], ext_map)

                    for field_key, field_info in result["fields"].items():
                        translated_value = field_info["translated"]
                        original_text = get_original_text(hotel, field_key)
                        if not original_text or not original_text.strip():
                            continue

                        db.add(
                            TranslationHistory(
                                source_text=original_text,
                                translated_text=translated_value,
                                source_lang="zh",
                                target_lang="en",
                                translation_type=translation_type,
                                reference_used=False,
                                glossary_used=False,
                                confidence_score=None,
                                review_status=ReviewStatus.PENDING,
                                booking_reference=None,
                                operator_name="translate_cli",
                            )
                        )

                    # 按计划要求：每个酒店单独 commit，确保中途失败不会回滚已成功的
                    await db.commit()
                    succeeded_hotels.append(result["hotel_id"])
                except Exception as e:
                    await db.rollback()
                    failed_hotels.append((result["hotel_id"], str(e)))

            console.print(
                f"\n[bold]Result:[/bold] [green]✓ {len(succeeded_hotels)} succeeded[/green], "
                f"[red]✗ {len(failed_hotels)} failed[/red]"
            )
            if failed_hotels:
                console.print("[red]Failed hotels:[/red]")
                for hid, err in failed_hotels:
                    console.print(f"  - {hid}: {err}")
            if succeeded_hotels:
                console.print("[green]✓ Translations applied successfully![/green]")

        else:
            console.print("\n[blue]Dry run - no changes made to database.[/blue]")


# ===================================================================
# 5 个查询子命令
# ===================================================================


@app.command(name="by-id")
def by_id(
    hotel_id: str = typer.Argument(..., help="Hotel UUID to translate"),
    dry_run: bool = DRY_RUN_OPTION,
    no_ai: bool = NO_AI_OPTION,
    concurrency: int = CONCURRENCY_OPTION,
    export_csv: Optional[Path] = EXPORT_CSV_OPTION,
    export_excel: Optional[Path] = EXPORT_EXCEL_OPTION,
):
    """Translate a single hotel by its UUID."""

    async def run() -> None:
        async with get_db_context() as db:
            stmt = (
                select(Hotel)
                .options(selectinload(Hotel.rooms))
                .where(Hotel.id == hotel_id)
            )
            result = await db.execute(stmt)
            hotel = result.scalar_one_or_none()

            if not hotel:
                console.print(f"[red]Hotel not found: {hotel_id}[/red]")
                raise typer.Exit(1)

            await _translate_workflow(
                hotel_ids=[hotel_id],
                hotels_for_display=[hotel],
                dry_run=dry_run,
                no_ai=no_ai,
                concurrency=concurrency,
                export_csv=export_csv,
                export_excel=export_excel,
            )

    asyncio.run(run())


@app.command(name="by-search")
def by_search(
    keyword: str = typer.Argument(
        ..., help="Search keyword (matches name_cn or name_en)"
    ),
    dry_run: bool = DRY_RUN_OPTION,
    no_ai: bool = NO_AI_OPTION,
    concurrency: int = CONCURRENCY_OPTION,
    export_csv: Optional[Path] = EXPORT_CSV_OPTION,
    export_excel: Optional[Path] = EXPORT_EXCEL_OPTION,
):
    """Search and translate hotels by keyword (partial match on name_cn or name_en)."""

    async def run() -> None:
        async with get_db_context() as db:
            stmt = (
                select(Hotel)
                .options(selectinload(Hotel.rooms))
                .where(
                    (Hotel.name_cn.ilike(f"%{keyword}%"))
                    | (Hotel.name_en.ilike(f"%{keyword}%"))
                )
                .order_by(Hotel.updated_at.desc())
            )
            result = await db.execute(stmt)
            hotels = list(result.scalars().all())

            if not hotels:
                console.print(f"[red]No hotels found matching: {keyword}[/red]")
                raise typer.Exit(1)

            console.print(
                f"[green]Found {len(hotels)} hotel(s) matching '{keyword}'[/green]"
            )
            await _translate_workflow(
                hotel_ids=[str(h.id) for h in hotels],
                hotels_for_display=hotels,
                dry_run=dry_run,
                no_ai=no_ai,
                concurrency=concurrency,
                export_csv=export_csv,
                export_excel=export_excel,
            )

    asyncio.run(run())


@app.command(name="by-brand")
def by_brand(
    brand: HotelBrand = typer.Argument(
        ..., help="Hotel brand", case_sensitive=False
    ),
    dry_run: bool = DRY_RUN_OPTION,
    no_ai: bool = NO_AI_OPTION,
    concurrency: int = CONCURRENCY_OPTION,
    export_csv: Optional[Path] = EXPORT_CSV_OPTION,
    export_excel: Optional[Path] = EXPORT_EXCEL_OPTION,
):
    """Translate hotels filtered by brand (atour / atour_x / zhotel / ahaus)."""

    async def run() -> None:
        async with get_db_context() as db:
            stmt = (
                select(Hotel)
                .options(selectinload(Hotel.rooms))
                .where(Hotel.brand == brand)
                .order_by(Hotel.updated_at.desc())
            )
            result = await db.execute(stmt)
            hotels = list(result.scalars().all())

            if not hotels:
                console.print(f"[red]No hotels found for brand: {brand.value}[/red]")
                raise typer.Exit(1)

            console.print(
                f"[green]Found {len(hotels)} hotel(s) for brand '{brand.value}'[/green]"
            )
            await _translate_workflow(
                hotel_ids=[str(h.id) for h in hotels],
                hotels_for_display=hotels,
                dry_run=dry_run,
                no_ai=no_ai,
                concurrency=concurrency,
                export_csv=export_csv,
                export_excel=export_excel,
            )

    asyncio.run(run())


@app.command(name="by-filter")
def by_filter(
    brand: Optional[HotelBrand] = typer.Option(
        None, "--brand", help="Filter by brand"
    ),
    city: Optional[str] = typer.Option(None, "--city", help="Filter by city"),
    country: Optional[str] = typer.Option(
        None, "--country", help="Filter by country code (e.g., CN)"
    ),
    status: Optional[str] = typer.Option(
        None,
        "--status",
        help="Filter by status (draft / pending_review / approved / published / suspended)",
    ),
    dry_run: bool = DRY_RUN_OPTION,
    no_ai: bool = NO_AI_OPTION,
    concurrency: int = CONCURRENCY_OPTION,
    export_csv: Optional[Path] = EXPORT_CSV_OPTION,
    export_excel: Optional[Path] = EXPORT_EXCEL_OPTION,
):
    """Translate hotels by flexible filter conditions (brand, city, country, status).

    At least one filter must be specified.
    """

    async def run() -> None:
        async with get_db_context() as db:
            stmt = select(Hotel).options(selectinload(Hotel.rooms))

            conditions = []
            filter_desc_parts = []

            if brand is not None:
                conditions.append(Hotel.brand == brand)
                filter_desc_parts.append(f"brand={brand.value}")
            if city is not None:
                conditions.append(Hotel.city == city)
                filter_desc_parts.append(f"city={city}")
            if country is not None:
                conditions.append(Hotel.country_code == country)
                filter_desc_parts.append(f"country={country}")
            if status is not None:
                try:
                    status_enum = HotelStatus(status)
                    conditions.append(Hotel.status == status_enum)
                    filter_desc_parts.append(f"status={status}")
                except ValueError:
                    valid = ", ".join(s.value for s in HotelStatus)
                    console.print(
                        f"[red]Invalid status: '{status}'. Valid values: {valid}[/red]"
                    )
                    raise typer.Exit(1)

            if not conditions:
                console.print(
                    "[red]At least one filter is required "
                    "(--brand, --city, --country, --status)[/red]"
                )
                raise typer.Exit(1)

            for cond in conditions:
                stmt = stmt.where(cond)

            stmt = stmt.order_by(Hotel.updated_at.desc())

            result = await db.execute(stmt)
            hotels = list(result.scalars().all())

            if not hotels:
                desc = ", ".join(filter_desc_parts)
                console.print(f"[red]No hotels found for filter: {desc}[/red]")
                raise typer.Exit(1)

            console.print(
                f"[green]Found {len(hotels)} hotel(s) matching filters[/green]"
            )
            await _translate_workflow(
                hotel_ids=[str(h.id) for h in hotels],
                hotels_for_display=hotels,
                dry_run=dry_run,
                no_ai=no_ai,
                concurrency=concurrency,
                export_csv=export_csv,
                export_excel=export_excel,
            )

    asyncio.run(run())


@app.command(name="all-untranslated")
def all_untranslated(
    dry_run: bool = DRY_RUN_OPTION,
    no_ai: bool = NO_AI_OPTION,
    concurrency: int = CONCURRENCY_OPTION,
    export_csv: Optional[Path] = EXPORT_CSV_OPTION,
    export_excel: Optional[Path] = EXPORT_EXCEL_OPTION,
):
    """Translate all hotels where name_en IS NULL (missing English name)."""

    async def run() -> None:
        async with get_db_context() as db:
            stmt = (
                select(Hotel)
                .options(selectinload(Hotel.rooms))
                .where(Hotel.name_en.is_(None))
                .order_by(Hotel.updated_at.desc())
            )
            result = await db.execute(stmt)
            hotels = list(result.scalars().all())

            if not hotels:
                console.print(
                    "[green]All hotels already have English names. Nothing to translate.[/green]"
                )
                raise typer.Exit(0)

            console.print(
                f"[green]Found {len(hotels)} hotel(s) with missing English name[/green]"
            )
            await _translate_workflow(
                hotel_ids=[str(h.id) for h in hotels],
                hotels_for_display=hotels,
                dry_run=dry_run,
                no_ai=no_ai,
                concurrency=concurrency,
                export_csv=export_csv,
                export_excel=export_excel,
            )

    asyncio.run(run())


if __name__ == "__main__":
    app()
