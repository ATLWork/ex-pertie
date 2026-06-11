"""
Integration tests for translate_cli.py CLI tool.

Tests cover all 5 commands (by-id, by-search, by-brand, by-filter, all-untranslated),
CSV/Excel export, and error handling including partial failure recovery.

Strategy:
- Mock ``BatchHotelTranslator.translate_batch`` to return controlled test results
- Mock ``get_db_context`` to provide a test async session (in-memory SQLite)
- Use ``typer.testing.CliRunner`` for synchronous CLI invocation
"""

import asyncio
import csv
import os
import sys
import tempfile
from contextlib import asynccontextmanager
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from typer.testing import CliRunner

# Ensure backend/ is on the path so scripts.translate_cli imports work
sys.path.insert(0, str(Path(__file__).parent.parent))

# ---------------------------------------------------------------------------
# Module-level test DB engine (shared across all tests in this file)
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def _cli_test_engine():
    """Module-scoped async engine on in-memory SQLite for CLI tests."""
    from app.core.database import Base

    engine = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False)

    async def _create_tables():
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

    asyncio.run(_create_tables())
    yield engine

    async def _drop_tables():
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.drop_all)

    asyncio.run(_drop_tables())
    try:
        asyncio.run(engine.dispose())
    except Exception:
        pass


def _make_session_factory(engine):
    """Create an async sessionmaker bound to *engine*."""
    return async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)


# ---------------------------------------------------------------------------
# Helper: build a mock ``get_db_context`` that yields a session from the
# module-level test engine.  Must return an *async context manager* because
# the real ``get_db_context`` is decorated with ``@asynccontextmanager``.
# ---------------------------------------------------------------------------


def _make_mock_db_context(session_factory):
    """Return a mock ``get_db_context`` replacement."""

    @asynccontextmanager
    async def _mock():
        async with session_factory() as session:
            yield session

    return _mock


# ---------------------------------------------------------------------------
# Helpers: create test hotels in the test DB
# ---------------------------------------------------------------------------


async def _seed_hotel(session: AsyncSession, **kwargs) -> dict:
    """Insert a single hotel and return its id + name_cn."""
    from app.models.hotel import Hotel, HotelBrand, HotelStatus

    defaults = {
        "name_cn": "测试酒店",
        "name_en": None,
        "brand": HotelBrand.ATour,
        "status": HotelStatus.DRAFT,
        "country_code": "CN",
        "province": "上海市",
        "city": "上海",
        "district": "浦东新区",
        "address_cn": "浦东新区某路123号",
        "postal_code": "200000",
    }
    defaults.update(kwargs)
    hotel = Hotel(**defaults)
    session.add(hotel)
    await session.flush()
    await session.refresh(hotel)
    return {"id": str(hotel.id), "name_cn": hotel.name_cn}


async def _truncate_hotels(session: AsyncSession) -> None:
    """Remove all hotels (and cascaded rooms/extensions) from the test DB."""
    from app.models.hotel import Hotel
    from app.models.translation import TranslationHistory

    from sqlalchemy import delete

    await session.execute(delete(TranslationHistory))
    await session.execute(delete(Hotel))
    await session.commit()


# ---------------------------------------------------------------------------
# Mock translate_batch result builder
# ---------------------------------------------------------------------------


def _make_translate_result(hotel_id: str, fields: dict = None, errors: list = None) -> dict:
    """Build a result dict matching BatchHotelTranslator.translate_hotel output.

    Each field value must be a dict with ``translated``, ``source``, and ``level`` keys.
    """
    return {
        "hotel_id": hotel_id,
        "fields": fields or {"name_en": {"translated": "Test Hotel EN", "source": "CACHE", "level": "hotel"}},
        "errors": errors or [],
    }


# ===================================================================
# 14 test cases
# ===================================================================

# ---------------------------------------------------------------------------
# 1. test_by_id_help
# ---------------------------------------------------------------------------


def test_by_id_help():
    """Verify ``by-id --help`` prints correct usage info."""
    from scripts.translate_cli import app

    runner = CliRunner()
    result = runner.invoke(app, ["by-id", "--help"])
    assert result.exit_code == 0
    assert "UUID" in result.output or "HOTEL" in result.output.upper()


# ---------------------------------------------------------------------------
# 2. test_by_id_not_found
# ---------------------------------------------------------------------------


def test_by_id_not_found(_cli_test_engine):
    """Invalid UUID → exit code 1 with 'not found' message."""
    from scripts.translate_cli import app

    session_factory = _make_session_factory(_cli_test_engine)
    mock_db_ctx = _make_mock_db_context(session_factory)

    with patch("scripts.translate_cli.get_db_context", side_effect=mock_db_ctx):
        runner = CliRunner()
        result = runner.invoke(app, ["by-id", "00000000-0000-0000-0000-000000000000"])
        assert result.exit_code == 1
        assert "not found" in result.output.lower()


# ---------------------------------------------------------------------------
# 3. test_by_id_dry_run
# ---------------------------------------------------------------------------


def test_by_id_dry_run(_cli_test_engine):
    """``--dry-run`` flag does not write to DB."""
    from scripts.translate_cli import app

    session_factory = _make_session_factory(_cli_test_engine)
    mock_db_ctx = _make_mock_db_context(session_factory)

    # Seed a hotel
    async def _seed():
        async with session_factory() as session:
            h = await _seed_hotel(session, name_cn="DryRun酒店")
            await session.commit()
            return h

    hotel = asyncio.run(_seed())

    mock_result = _make_translate_result(hotel["id"], {"name_en": {"translated": "Dry Run Hotel", "source": "CACHE", "level": "hotel"}})
    mock_translator = MagicMock()
    mock_translator.translate_batch = AsyncMock(return_value=[mock_result])

    with (
        patch("scripts.translate_cli.get_db_context", side_effect=mock_db_ctx),
        patch("scripts.translate_cli.BatchHotelTranslator", return_value=mock_translator),
    ):
        runner = CliRunner()
        result = runner.invoke(app, ["by-id", hotel["id"], "--dry-run", "--no-ai"])
        assert result.exit_code == 0
        assert "Dry run" in result.output

    # Verify name_en was NOT updated (dry run)
    async def _check():
        async with session_factory() as session:
            from app.models.hotel import Hotel

            stmt = select(Hotel).where(Hotel.id == hotel["id"])
            r = await session.execute(stmt)
            h = r.scalar_one()
            assert h.name_en is None

    asyncio.run(_check())


# ---------------------------------------------------------------------------
# 4. test_by_id_confirm_write
# ---------------------------------------------------------------------------


def test_by_id_confirm_write(_cli_test_engine):
    """After confirmation, DB fields are updated and TranslationHistory created."""
    from scripts.translate_cli import app

    session_factory = _make_session_factory(_cli_test_engine)
    mock_db_ctx = _make_mock_db_context(session_factory)

    # Seed
    async def _seed():
        async with session_factory() as session:
            h = await _seed_hotel(session, name_cn="写入测试酒店")
            await session.commit()
            return h

    hotel = asyncio.run(_seed())

    mock_result = _make_translate_result(hotel["id"], {"name_en": {"translated": "Write Test Hotel", "source": "MACHINE", "level": "hotel"}})
    mock_translator = MagicMock()
    mock_translator.translate_batch = AsyncMock(return_value=[mock_result])

    with (
        patch("scripts.translate_cli.get_db_context", side_effect=mock_db_ctx),
        patch("scripts.translate_cli.BatchHotelTranslator", return_value=mock_translator),
    ):
        runner = CliRunner()
        result = runner.invoke(app, ["by-id", hotel["id"], "--no-ai"], input="y\n")
        assert result.exit_code == 0

    # Verify DB write
    async def _verify():
        async with session_factory() as session:
            from app.models.hotel import Hotel
            from app.models.translation import TranslationHistory

            stmt = select(Hotel).where(Hotel.id == hotel["id"])
            r = await session.execute(stmt)
            h = r.scalar_one()
            assert h.name_en == "Write Test Hotel"

            hist_stmt = select(TranslationHistory).where(
                TranslationHistory.operator_name == "translate_cli"
            )
            hist_r = await session.execute(hist_stmt)
            histories = hist_r.scalars().all()
            assert len(histories) >= 1

    asyncio.run(_verify())


# ---------------------------------------------------------------------------
# 5. test_by_search_found
# ---------------------------------------------------------------------------


def test_by_search_found(_cli_test_engine):
    """Search keyword returns matching hotels and displays preview."""
    from scripts.translate_cli import app

    session_factory = _make_session_factory(_cli_test_engine)
    mock_db_ctx = _make_mock_db_context(session_factory)

    async def _seed():
        async with session_factory() as session:
            h1 = await _seed_hotel(session, name_cn="上海亚朵酒店")
            h2 = await _seed_hotel(session, name_cn="北京亚朵酒店")
            await session.commit()
            return [h1, h2]

    hotels = asyncio.run(_seed())
    mock_results = [
        _make_translate_result(h["id"], {"name_en": {"translated": f"Hotel {h['name_cn']} EN", "source": "CACHE", "level": "hotel"}})
        for h in hotels
    ]

    mock_translator = MagicMock()
    mock_translator.translate_batch = AsyncMock(return_value=mock_results)

    with (
        patch("scripts.translate_cli.get_db_context", side_effect=mock_db_ctx),
        patch("scripts.translate_cli.BatchHotelTranslator", return_value=mock_translator),
    ):
        runner = CliRunner()
        result = runner.invoke(app, ["by-search", "亚朵", "--dry-run", "--no-ai"])
        assert result.exit_code == 0
        assert "Found 2 hotel" in result.output


# ---------------------------------------------------------------------------
# 6. test_by_search_not_found
# ---------------------------------------------------------------------------


def test_by_search_not_found(_cli_test_engine):
    """No matching hotels → exit code 1."""
    from scripts.translate_cli import app

    session_factory = _make_session_factory(_cli_test_engine)
    mock_db_ctx = _make_mock_db_context(session_factory)

    with patch("scripts.translate_cli.get_db_context", side_effect=mock_db_ctx):
        runner = CliRunner()
        result = runner.invoke(app, ["by-search", "ZZZ_NONEXISTENT_ZZZ"])
        assert result.exit_code == 1
        assert "No hotels found" in result.output


# ---------------------------------------------------------------------------
# 7. test_by_brand_valid
# ---------------------------------------------------------------------------


def test_by_brand_valid(_cli_test_engine):
    """Valid brand returns matching hotels."""
    from scripts.translate_cli import app

    session_factory = _make_session_factory(_cli_test_engine)
    mock_db_ctx = _make_mock_db_context(session_factory)

    async def _seed():
        async with session_factory() as session:
            from app.models.hotel import HotelBrand

            h = await _seed_hotel(session, name_cn="亚朵X酒店", brand=HotelBrand.ATourX)
            await session.commit()
            return h

    hotel = asyncio.run(_seed())
    mock_results = [_make_translate_result(hotel["id"])]

    mock_translator = MagicMock()
    mock_translator.translate_batch = AsyncMock(return_value=mock_results)

    with (
        patch("scripts.translate_cli.get_db_context", side_effect=mock_db_ctx),
        patch("scripts.translate_cli.BatchHotelTranslator", return_value=mock_translator),
    ):
        runner = CliRunner()
        result = runner.invoke(app, ["by-brand", "atour_x", "--dry-run", "--no-ai"])
        assert result.exit_code == 0
        assert "atour_x" in result.output


# ---------------------------------------------------------------------------
# 8. test_by_brand_invalid
# ---------------------------------------------------------------------------


def test_by_brand_invalid():
    """Invalid brand → exit code != 0 (Typer argument parsing error)."""
    from scripts.translate_cli import app

    runner = CliRunner()
    result = runner.invoke(app, ["by-brand", "invalid_brand_name"])
    assert result.exit_code != 0


# ---------------------------------------------------------------------------
# 9. test_by_filter_multi
# ---------------------------------------------------------------------------


def test_by_filter_multi(_cli_test_engine):
    """Combined filter (brand + city) finds matching hotels."""
    from scripts.translate_cli import app

    session_factory = _make_session_factory(_cli_test_engine)
    mock_db_ctx = _make_mock_db_context(session_factory)

    async def _seed():
        async with session_factory() as session:
            from app.models.hotel import HotelBrand

            await _truncate_hotels(session)
            h1 = await _seed_hotel(session, name_cn="上海亚朵A", brand=HotelBrand.ATour, city="上海")
            await _seed_hotel(session, name_cn="北京亚朵B", brand=HotelBrand.ATour, city="北京")
            await session.commit()
            return h1

    hotel = asyncio.run(_seed())
    mock_results = [_make_translate_result(hotel["id"])]

    mock_translator = MagicMock()
    mock_translator.translate_batch = AsyncMock(return_value=mock_results)

    with (
        patch("scripts.translate_cli.get_db_context", side_effect=mock_db_ctx),
        patch("scripts.translate_cli.BatchHotelTranslator", return_value=mock_translator),
    ):
        runner = CliRunner()
        result = runner.invoke(
            app, ["by-filter", "--brand", "atour", "--city", "上海", "--dry-run", "--no-ai"]
        )
        assert result.exit_code == 0
        assert "Found 1 hotel" in result.output


# ---------------------------------------------------------------------------
# 10. test_by_filter_no_params
# ---------------------------------------------------------------------------


def test_by_filter_no_params():
    """No filter parameters → exit code 1."""
    from scripts.translate_cli import app

    runner = CliRunner()
    result = runner.invoke(app, ["by-filter"])
    assert result.exit_code == 1
    assert "At least one filter" in result.output


# ---------------------------------------------------------------------------
# 11. test_all_untranslated
# ---------------------------------------------------------------------------


def test_all_untranslated(_cli_test_engine):
    """Query hotels where name_en IS NULL."""
    from scripts.translate_cli import app

    session_factory = _make_session_factory(_cli_test_engine)
    mock_db_ctx = _make_mock_db_context(session_factory)

    async def _seed():
        async with session_factory() as session:
            h1 = await _seed_hotel(session, name_cn="未翻译酒店", name_en=None)
            await _seed_hotel(session, name_cn="已翻译酒店", name_en="Already EN")
            await session.commit()
            return h1

    hotel = asyncio.run(_seed())
    mock_results = [_make_translate_result(hotel["id"])]

    mock_translator = MagicMock()
    mock_translator.translate_batch = AsyncMock(return_value=mock_results)

    with (
        patch("scripts.translate_cli.get_db_context", side_effect=mock_db_ctx),
        patch("scripts.translate_cli.BatchHotelTranslator", return_value=mock_translator),
    ):
        runner = CliRunner()
        result = runner.invoke(app, ["all-untranslated", "--dry-run", "--no-ai"])
        assert result.exit_code == 0
        assert "1 hotel" in result.output or "missing English" in result.output


# ---------------------------------------------------------------------------
# 12. test_csv_export
# ---------------------------------------------------------------------------


def test_csv_export(_cli_test_engine):
    """CSV export file contains 5-column header."""
    from scripts.translate_cli import app

    session_factory = _make_session_factory(_cli_test_engine)
    mock_db_ctx = _make_mock_db_context(session_factory)

    async def _seed():
        async with session_factory() as session:
            h = await _seed_hotel(session, name_cn="CSV导出酒店")
            await session.commit()
            return h

    hotel = asyncio.run(_seed())
    mock_results = [_make_translate_result(hotel["id"], {"name_en": {"translated": "CSV Export Hotel", "source": "CACHE", "level": "hotel"}})]

    mock_translator = MagicMock()
    mock_translator.translate_batch = AsyncMock(return_value=mock_results)

    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tmp:
        csv_path = tmp.name

    try:
        with (
            patch("scripts.translate_cli.get_db_context", side_effect=mock_db_ctx),
            patch("scripts.translate_cli.BatchHotelTranslator", return_value=mock_translator),
        ):
            runner = CliRunner()
            result = runner.invoke(
                app,
                ["by-id", hotel["id"], "--dry-run", "--no-ai", "--export-csv", csv_path],
            )
            assert result.exit_code == 0

        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader)
            assert header == ["Hotel ID", "Hotel Name", "Level", "Field", "Original", "Translated", "Source"]
            rows = list(reader)
            assert len(rows) >= 1
            assert rows[0][2] == "hotel"
            assert rows[0][6] == "CACHE"
    finally:
        os.unlink(csv_path)


# ---------------------------------------------------------------------------
# 13. test_excel_export
# ---------------------------------------------------------------------------


def test_excel_export(_cli_test_engine):
    """Excel export produces a valid .xlsx file."""
    pytest.importorskip("openpyxl", reason="openpyxl not installed")

    from scripts.translate_cli import app

    session_factory = _make_session_factory(_cli_test_engine)
    mock_db_ctx = _make_mock_db_context(session_factory)

    async def _seed():
        async with session_factory() as session:
            h = await _seed_hotel(session, name_cn="Excel导出酒店")
            await session.commit()
            return h

    hotel = asyncio.run(_seed())
    mock_results = [_make_translate_result(hotel["id"], {"name_en": {"translated": "Excel Export Hotel", "source": "AI_ENHANCED", "level": "hotel"}})]

    mock_translator = MagicMock()
    mock_translator.translate_batch = AsyncMock(return_value=mock_results)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        with (
            patch("scripts.translate_cli.get_db_context", side_effect=mock_db_ctx),
            patch("scripts.translate_cli.BatchHotelTranslator", return_value=mock_translator),
        ):
            runner = CliRunner()
            result = runner.invoke(
                app,
                ["by-id", hotel["id"], "--dry-run", "--no-ai", "--export-excel", xlsx_path],
            )
            assert result.exit_code == 0

        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        assert ws.title == "Translations"
        header = [cell.value for cell in ws[1]]
        assert header == ["Hotel ID", "Hotel Name", "Level", "Field", "Original", "Translated", "Source"]
        assert ws.cell(row=2, column=1).value is not None
    finally:
        os.unlink(xlsx_path)


# ---------------------------------------------------------------------------
# 14. test_partial_failure_recovery
# ---------------------------------------------------------------------------


def test_partial_failure_recovery(_cli_test_engine):
    """One hotel failing translation does not block others."""
    from scripts.translate_cli import app

    session_factory = _make_session_factory(_cli_test_engine)
    mock_db_ctx = _make_mock_db_context(session_factory)

    async def _seed():
        async with session_factory() as session:
            h1 = await _seed_hotel(session, name_cn="成功酒店")
            h2 = await _seed_hotel(session, name_cn="失败酒店")
            await session.commit()
            return h1, h2

    h1, h2 = asyncio.run(_seed())
    mock_results = [
        _make_translate_result(h1["id"], {"name_en": {"translated": "Success Hotel", "source": "CACHE", "level": "hotel"}}),
        _make_translate_result(h2["id"], fields={}, errors=["Translation failed for name_en"]),
    ]

    mock_translator = MagicMock()
    mock_translator.translate_batch = AsyncMock(return_value=mock_results)

    with (
        patch("scripts.translate_cli.get_db_context", side_effect=mock_db_ctx),
        patch("scripts.translate_cli.BatchHotelTranslator", return_value=mock_translator),
    ):
        runner = CliRunner()
        result = runner.invoke(app, ["by-search", "酒店", "--dry-run", "--no-ai"])
        assert result.exit_code == 0
        assert "Success Hotel" in result.output
        assert "1 errors" in result.output or "error" in result.output.lower()
