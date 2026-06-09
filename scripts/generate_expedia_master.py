import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

FILE = 'reference/Expedia 亚朵集团酒店上线表格 - APM Onboarding Sheet example.xlsx'

# ── 1. 读取数据 ─────────────────────────────────────────────────────────────
store = pd.read_excel(FILE, sheet_name='门店基础信息', header=0)
rooms = pd.read_excel(FILE, sheet_name='亚朵房型信息', header=0)

store_cols = {
    '门店ID': 'Hotel Code',
    '门店名称': 'Hotel Name',
    '品牌': 'Brand',
    '所属城市': 'City',
    '门店地址': 'Address',
    '电话': 'Phone Raw',
    '邮编': 'Postal Code',
}
rooms_cols = {
    '集团酒店编号': 'Hotel Code',
    '房型名称': 'Room Type Name CN',
    '集团房型代码': 'Room Type Code',
    '楼层': 'Floor',
    '房型面积': 'Room Size',
    '可容纳人数': 'Max Occupancy',
    '床型': 'Bed Type Raw',
    '床个数': 'Bed Count',
    '该房型的房间数': 'Room Count',
    '是否有窗': 'Window',
    '是否有会议室': 'Meeting Room',
}

store_df = store[list(store_cols.keys())].rename(columns=store_cols)
rooms_df = rooms[list(rooms_cols.keys())].rename(columns=rooms_cols)

# ── 2. 电话解析 ──────────────────────────────────────────────────────────────
def parse_phone(raw):
    if pd.isna(raw):
        return '', '', ''
    s = str(raw).strip()
    m = re.match(r'0?(\d{2,4})-(\d{7,8})', s)
    if m:
        area = m.group(1)
        num = m.group(2)
        country = '86'
        if area == '10':
            country = '86'
        elif len(area) == 3:
            country = '86'
        return country, area, num
    m2 = re.match(r'(\d{2,3})-(\d{7,8})', s)
    if m2:
        return '86', m2.group(1), m2.group(2)
    return '86', '', s

phone_parsed = store_df['Phone Raw'].apply(parse_phone)
store_df['Phone Country'] = phone_parsed.apply(lambda x: x[0])
store_df['Phone Area'] = phone_parsed.apply(lambda x: x[1])
store_df['Phone Number'] = phone_parsed.apply(lambda x: x[2])
store_df.drop(columns=['Phone Raw'], inplace=True)

# ── 3. 床型映射 ─────────────────────────────────────────────────────────────
def map_bed(bed_raw):
    if pd.isna(bed_raw):
        return '1 QueenBed'
    s = str(bed_raw).strip()
    # 沙发床特殊处理
    if '沙发床' in s:
        if '1.8' in s or '2' in s:
            return '1 KingBed'
        return '1 QueenBed'
    # 1.8 / 2.0 / 2 → KingBed
    if re.match(r'^2(\.0?)?$', s) or s == '1.8' or s == '2':
        return '1 KingBed'
    # 1.5 / 1.5,1.5 → QueenBed
    if s == '1.5' or s == '1.5,1.5':
        return '1 QueenBed'
    # 1.2,1.2 → 2 TwinBed
    if s == '1.2,1.2':
        return '2 TwinBed'
    # 1.2,1.5 / 1.5,1.2 → 1 QueenBed & 1 TwinBed
    if s == '1.2,1.5' or s == '1.5,1.2':
        return '1 QueenBed&1 TwinBed'
    # 1.8,1.2 / 2,1.2 → 1 KingBed & 1 TwinBed
    if s == '1.8,1.2' or s == '2,1.2':
        return '1 KingBed&1 TwinBed'
    # 1.8,1.0 → 1 KingBed
    if s == '1.8,1.0':
        return '1 KingBed'
    # 1.35,1.35 → 2 QueenBed
    if s == '1.35,1.35':
        return '2 QueenBed'
    # 1.5,1.5 → 2 QueenBed
    if s == '1.5,1.5':
        return '2 QueenBed'
    # 默认 QueenBed
    return '1 QueenBed'

rooms_df['BeddingOption1'] = rooms_df['Bed Type Raw'].apply(map_bed)

# ── 4. 生成 Sheet1 ───────────────────────────────────────────────────────────
hotel_s1 = store_df.copy()
hotel_s1['Country'] = 'China'
hotel_s1['Time Zone'] = 'China Standard Time'
hotel_s1['Check-in Time'] = '14:00:00'
hotel_s1['Check-out Time'] = '12:00:00'
hotel_s1['Currency'] = 'CNY'
hotel_s1['Business Model'] = 'Dual'
hotel_s1['Language Preference'] = 'Simplified Chinese'
hotel_s1['Rate Acquisition Type'] = 'SellLAR'
hotel_s1['Min Adult Age'] = 18
hotel_s1['Rates Inclusive of Taxes'] = 'Yes'
hotel_s1['Cancellation Time'] = '12:00:00'
hotel_s1['Photos'] = 'Excel sheet (with media links)'
hotel_s1['Structure Type'] = 'Hotel'

s1_cols = [
    'Hotel Code', 'Hotel Name', 'Brand', 'City', 'Country',
    'Address', 'Phone Country', 'Phone Area', 'Phone Number',
    'Postal Code', 'Time Zone', 'Check-in Time', 'Check-out Time',
    'Currency', 'Business Model', 'Language Preference',
    'Rate Acquisition Type', 'Min Adult Age', 'Rates Inclusive of Taxes',
    'Cancellation Time', 'Photos', 'Structure Type'
]
hotel_s1 = hotel_s1[[c for c in s1_cols if c in hotel_s1.columns]]

# ── 5. 生成 Sheet2 ────────────────────────────────────────────────────────────
rooms_df['Max Adults'] = rooms_df['Max Occupancy']
rooms_df['Max Children'] = 0
rooms_df['Smoking'] = 'NonSmoking'
rooms_df['RNS Flag'] = 'Yes'
rooms_df['Rate Plan Type'] = 'Standalone'
rooms_df['Value Add 1'] = 'Free Wireless Internet'
rooms_df['Cancellation Window Hours'] = 18
rooms_df['Outside Window Penalty'] = 'Full Cost of Stay'
rooms_df['Inside Window Penalty'] = 'Full Cost of Stay'
rooms_df['Min Advance Booking Days'] = 0
rooms_df['Max Advance Booking Days'] = 180
rooms_df['Min Length of Stay'] = 1
rooms_df['Max Length of Stay'] = 28
rooms_df['Waive Taxes Enabled'] = 'Yes'

def make_rate_plan(row, model):
    code = row['Room Type Code']
    if model == 'Agency':
        return f'{code}-HC-A', 'Room Only HC A', 'Agency'
    else:
        return f'{code}-EC-A', 'Room Only EC A', 'Merchant'

agencies = rooms_df.copy()
agencies[['Rate Plan Code', 'Rate Plan Name', 'Rate Plan Business Model']] = agencies.apply(
    lambda r: pd.Series(make_rate_plan(r, 'Agency')), axis=1
)

merchants = rooms_df.copy()
merchants[['Rate Plan Code', 'Rate Plan Name', 'Rate Plan Business Model']] = merchants.apply(
    lambda r: pd.Series(make_rate_plan(r, 'Merchant')), axis=1
)

rooms_s2 = pd.concat([agencies, merchants], ignore_index=True)

s2_cols = [
    'Hotel Code', 'Room Type Code', 'Room Type Name CN', 'Max Occupancy',
    'Max Adults', 'Max Children', 'BeddingOption1', 'Smoking', 'RNS Flag',
    'Rate Plan Code', 'Rate Plan Name', 'Rate Plan Type',
    'Rate Plan Business Model', 'Value Add 1',
    'Cancellation Window Hours', 'Outside Window Penalty', 'Inside Window Penalty',
    'Min Advance Booking Days', 'Max Advance Booking Days',
    'Min Length of Stay', 'Max Length of Stay', 'Waive Taxes Enabled'
]
rooms_s2 = rooms_s2[[c for c in s2_cols if c in rooms_s2.columns]]

# ── 6. 写入 Excel ─────────────────────────────────────────────────────────────
out_file = f'output/酒店主数据导入_{datetime.now().strftime("%Y-%m-%d")}.xlsx'

wb = Workbook()
ws1 = wb.active
ws1.title = 'HotelMaster'

# 样式
header_fill = PatternFill('solid', start_color='366092')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_sheet(ws, df, sheet_name_cn):
    # 标题行
    ws.cell(row=1, column=1, value=f'{sheet_name_cn} - 生成日期: {datetime.now().strftime("%Y-%m-%d")}')
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    # 表头
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # 数据
    for row_idx, row in df.iterrows():
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx + 3, column=col_idx, value=val if pd.notna(val) else '')
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    # 列宽
    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

write_sheet(ws1, hotel_s1, '酒店主数据')

ws2 = wb.create_sheet('RoomMaster')
write_sheet(ws2, rooms_s2, '房间主数据')

wb.save(out_file)
print(f'已生成: {out_file}')
print(f'HotelMaster: {len(hotel_s1)} 家酒店')
print(f'RoomMaster: {len(rooms_s2)} 条房型记录 (每房型 2 行: Agency + Merchant)')